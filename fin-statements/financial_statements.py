#!/usr/bin/env python3
"""
Tally Financial Statements Generator
Reads a Tally SQLite export and produces:
  • Schedule III Balance Sheet
  • Statement of Profit & Loss
  • 3-Year Projected P&L and Balance Sheet (banker/investor format)

Run:  python financial_statements.py
Requires: openpyxl  (pip install openpyxl)
"""
import os
import sqlite3
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from datetime import date
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# ─── Schedule III primary-group classification ────────────────────────────────
# Values: (side, schedule_head, is_asset)
#   side:  'equity' | 'noncurrent_liab' | 'current_liab' | 'noncurrent_asset' | 'current_asset'
#   schedule_head: display label used in grouping
#   is_asset: True if the natural balance is debit (asset)
#
# Tally sign convention in closing_balance:
#   negative → debit balance   (normal for assets / expenses)
#   positive → credit balance  (normal for liabilities / income)
#
# For display we always show amounts as POSITIVE on the face of statements:
#   asset display  =  – closing_balance   (negate the debit)
#   liab display   =  + closing_balance   (credit as-is)
#   revenue display=  + closing_balance   (credit as-is)
#   expense display=  – closing_balance   (negate the debit)

BS_MAP: dict[str, tuple[str, str, bool]] = {
    # ── Assets ───────────────────────────────────────────────────────────────
    "Fixed Assets":             ("noncurrent_asset", "Fixed Assets",                     True),
    "Investments":              ("noncurrent_asset", "Non-Current Investments",           True),
    "Deposits (Asset)":         ("noncurrent_asset", "Long-Term Loans & Advances",        True),
    "Loans & Advances (Asset)": ("noncurrent_asset", "Long-Term Loans & Advances",        True),
    "Misc. Expenses (ASSET)":   ("noncurrent_asset", "Other Non-Current Assets",          True),
    "Stock-in-hand":            ("current_asset",    "Inventories",                       True),
    "Sundry Debtors":           ("current_asset",    "Trade Receivables",                 True),
    "Cash-in-hand":             ("current_asset",    "Cash & Cash Equivalents",           True),
    "Bank Accounts":            ("current_asset",    "Cash & Cash Equivalents",           True),
    "Current Assets":           ("current_asset",    "Other Current Assets",              True),
    # ── Equity ───────────────────────────────────────────────────────────────
    "Capital Account":          ("equity",           "Share Capital",                     False),
    "Reserves & Surplus":       ("equity",           "Reserves & Surplus",                False),
    # ── Non-current liabilities ──────────────────────────────────────────────
    "Secured Loans":            ("noncurrent_liab",  "Long-Term Borrowings",              False),
    "Unsecured Loans":          ("noncurrent_liab",  "Long-Term Borrowings",              False),
    "Loans (Liability)":        ("noncurrent_liab",  "Long-Term Borrowings",              False),
    # ── Current liabilities ──────────────────────────────────────────────────
    "Bank OD A/c":              ("current_liab",     "Short-Term Borrowings",             False),
    "Sundry Creditors":         ("current_liab",     "Trade Payables",                    False),
    "Current Liabilities":      ("current_liab",     "Other Current Liabilities",         False),
    "Branch / Divisions":       ("current_liab",     "Other Current Liabilities",         False),
    "Duties & Taxes":           ("current_liab",     "Duties & Taxes (Net)",              False),
    "Provisions":               ("current_liab",     "Short-Term Provisions",             False),
    "Suspense A/c":             ("current_liab",     "Other Current Liabilities",         False),
}

PNL_MAP: dict[str, tuple[str, str]] = {
    "Sales Accounts":   ("revenue",  "Revenue from Operations"),
    "Direct Incomes":   ("revenue",  "Revenue from Operations"),
    "Indirect Incomes": ("revenue",  "Other Income"),
    "Purchase Accounts":("expense",  "Cost of Materials / Purchases"),
    "Direct Expenses":  ("expense",  "Direct Expenses"),
    "Indirect Expenses":("expense",  "Indirect Expenses"),
}

# Sub-groups of Indirect Expenses that are carved out on the face of P&L
FINANCE_COST_PARENTS = {"Finance Costs", "Interest & Late Filing Fees"}
EMPLOYEE_COST_PARENTS = {
    "Employee benefit expenses", "Contribution to Provident Funds & Others",
    "Salary", "Salaries", "Staff Salary",
}

# ─── Data classes ─────────────────────────────────────────────────────────────

@dataclass
class LedgerRow:
    name: str
    parent: str
    primary_group: str
    opening: float
    closing: float
    is_deemedpositive: bool

@dataclass
class PnlRow:
    """Aggregated P&L transaction total from trn_accounting."""
    primary_group: str
    parent: str           # Tally sub-group (e.g. "Finance Costs")
    total: float          # raw sum; negative = debit (expense/purchase), positive = credit (income)

@dataclass
class FinancialData:
    company: str
    period_from: str
    period_to: str
    period_label: str
    ledgers: list[LedgerRow]
    pnl_rows: list[PnlRow]    # from trn_accounting (kept for future use)
    pnl_balance: float         # P&L A/c closing (cumulative, used in Reserves & Surplus on BS)
    pnl_opening: float         # P&L A/c opening (prior year retained profit not yet in Reserves)
    opening_stock_override: float | None = None
    closing_stock_override: float | None = None

    # ── Balance-sheet ledger helpers ─────────────────────────────────────────

    def _ledgers_for(self, primary_group: str) -> list[LedgerRow]:
        return [l for l in self.ledgers if l.primary_group == primary_group]

    def _sum_closing(self, primary_group: str) -> float:
        return sum(l.closing for l in self._ledgers_for(primary_group))

    def _sum_opening(self, primary_group: str) -> float:
        return sum(l.opening for l in self._ledgers_for(primary_group))

    # ── P&L helpers (use mst_ledger closing_balance — matches Tally's own P&L) ──
    # trn_accounting totals include journal/inter-branch adjustments that Tally
    # excludes from its P&L computation. mst_ledger.closing_balance is the
    # authoritative figure Tally uses.
    # Sign convention for P&L ledgers:
    #   Revenue (Sales, Incomes): positive closing = credit = income as-is
    #   Expenses (Purchases, Costs): negative closing = debit = negate to display positive

    def _pnl_group_total(self, primary_group: str) -> float:
        """Sum of closing_balances for a primary group (matches Tally's P&L)."""
        return sum(l.closing for l in self._ledgers_for(primary_group))

    def _pnl_by_parent(self, primary_group: str) -> dict[str, float]:
        result: dict[str, float] = {}
        for l in self._ledgers_for(primary_group):
            result[l.parent] = result.get(l.parent, 0) + l.closing
        return result

    # ── Fixed assets note ─────────────────────────────────────────────────────

    def fixed_assets_schedule(self) -> list[dict]:
        by_subgroup: dict[str, dict] = {}
        for l in self._ledgers_for("Fixed Assets"):
            sg = l.parent
            if sg not in by_subgroup:
                by_subgroup[sg] = {"gross_open": 0, "gross_close": 0,
                                   "depr_open": 0, "depr_close": 0}
            d = by_subgroup[sg]
            is_depr = "depreciation" in l.name.lower() or "accumulated" in l.name.lower()
            if is_depr:
                d["depr_open"]  += l.opening   # credit = positive
                d["depr_close"] += l.closing
            else:
                d["gross_open"]  += -l.opening  # debit = negative → negate for display
                d["gross_close"] += -l.closing
        rows = []
        for sg, d in by_subgroup.items():
            rows.append({
                "name":        sg,
                "gross_open":  d["gross_open"],
                "additions":   max(0, d["gross_close"] - d["gross_open"]),
                "disposals":   max(0, d["gross_open"] - d["gross_close"]),
                "gross_close": d["gross_close"],
                "depr_open":   d["depr_open"],
                "depr_charge": max(0, d["depr_close"] - d["depr_open"]),
                "depr_close":  d["depr_close"],
                "net_open":    d["gross_open"]  - d["depr_open"],
                "net_close":   d["gross_close"] - d["depr_close"],
            })
        return sorted(rows, key=lambda r: r["name"])

    def net_fixed_assets(self) -> float:
        return sum(r["net_close"] for r in self.fixed_assets_schedule())

    # ── Stock ─────────────────────────────────────────────────────────────────

    def opening_stock(self) -> float:
        if self.opening_stock_override is not None:
            return self.opening_stock_override
        return -self._sum_opening("Stock-in-hand")   # debit balance → negate

    def closing_stock(self) -> float:
        if self.closing_stock_override is not None:
            return self.closing_stock_override
        return -self._sum_closing("Stock-in-hand")

    # ── P&L totals (from mst_ledger closing_balance) ─────────────────────────

    def revenue_from_ops(self) -> float:
        # Revenue accounts: credit nature → positive closing_balance → use as-is
        return (self._pnl_group_total("Sales Accounts")
                + self._pnl_group_total("Direct Incomes"))

    def other_income(self) -> float:
        return self._pnl_group_total("Indirect Incomes")

    def purchases(self) -> float:
        # Expense accounts: debit nature → negative closing_balance → negate to display positive
        return -self._pnl_group_total("Purchase Accounts")

    def direct_expenses(self) -> float:
        return -self._pnl_group_total("Direct Expenses")

    def finance_costs(self) -> float:
        ibd = self._pnl_by_parent("Indirect Expenses")
        return sum(-v for k, v in ibd.items() if k in FINANCE_COST_PARENTS)

    def employee_costs(self) -> float:
        ibd = self._pnl_by_parent("Indirect Expenses")
        return sum(-v for k, v in ibd.items() if k in EMPLOYEE_COST_PARENTS)

    def depreciation_from_fa(self) -> float:
        return sum(r["depr_charge"] for r in self.fixed_assets_schedule())

    def other_indirect_expenses(self) -> float:
        ibd = self._pnl_by_parent("Indirect Expenses")
        skip = FINANCE_COST_PARENTS | EMPLOYEE_COST_PARENTS
        return sum(-v for k, v in ibd.items() if k not in skip)

    def total_expenses(self) -> float:
        return (self.purchases()
                + self.direct_expenses()
                + self.employee_costs()
                + self.finance_costs()
                + self.depreciation_from_fa()
                + self.other_indirect_expenses()
                + self.opening_stock()
                - self.closing_stock())

    def profit_before_tax(self) -> float:
        return self.revenue_from_ops() + self.other_income() - self.total_expenses()

    def tax_expense(self) -> float:
        # Deferred Tax Liability ledger treatment: positive = credit = liability = tax expense
        dtl = next((l.closing for l in self.ledgers
                    if "deferred tax" in l.name.lower()), 0.0)
        return dtl  # simplified; actual tax computation needs I-T workings

    def profit_after_tax(self) -> float:
        return self.profit_before_tax() - self.tax_expense()

    # ── Balance sheet totals ──────────────────────────────────────────────────

    def share_capital(self) -> float:
        # Capital Account with EQUITY SHARE CAPITAL parent (exclude Reserve & Surplus ledger)
        total = 0.0
        for l in self._ledgers_for("Capital Account"):
            if "reserve" not in l.name.lower() and "profit" not in l.name.lower():
                total += l.closing   # credit = positive = equity
        return total

    def reserves_surplus(self) -> float:
        res = self._sum_closing("Reserves & Surplus")
        # Reserve & Surplus ledger under Capital Account
        for l in self._ledgers_for("Capital Account"):
            if "reserve" in l.name.lower():
                res += l.closing
        # Add current year P&L
        res += self.pnl_balance
        return res

    def long_term_borrowings(self) -> float:
        return (self._sum_closing("Secured Loans")
                + self._sum_closing("Unsecured Loans")
                + self._sum_closing("Loans (Liability)"))

    def short_term_borrowings(self) -> float:
        return self._sum_closing("Bank OD A/c")

    def trade_payables(self) -> float:
        return max(0, self._sum_closing("Sundry Creditors"))

    def duties_and_taxes_net(self) -> float:
        net = self._sum_closing("Duties & Taxes")
        # Positive net = payable (liability); negative net = refund (asset)
        return net

    def other_current_liabilities(self) -> float:
        # Duties & Taxes gets its own line; don't include it here
        cl     = self._sum_closing("Current Liabilities")
        branch = self._sum_closing("Branch / Divisions")
        susp   = self._sum_closing("Suspense A/c")
        ocl    = cl + max(0, branch) + susp
        # Remove Deferred Tax Liability from OCL (it's non-current)
        dtl = next((l.closing for l in self.ledgers
                    if "deferred tax" in l.name.lower()), 0.0)
        return max(0, ocl - dtl)

    def short_term_provisions(self) -> float:
        return max(0, self._sum_closing("Provisions"))

    def non_current_investments(self) -> float:
        return -self._sum_closing("Investments")

    def long_term_loans_advances(self) -> float:
        dep = -self._sum_closing("Deposits (Asset)")
        la  = -self._sum_closing("Loans & Advances (Asset)")
        return dep + la

    def other_noncurrent_assets(self) -> float:
        return -self._sum_closing("Misc. Expenses (ASSET)")

    def trade_receivables(self) -> float:
        return max(0, -self._sum_closing("Sundry Debtors"))

    def cash_and_bank(self) -> float:
        # Tally: for is_deemedpositive groups, negative closing = debit = asset (cash/bank balance)
        # Only count banks with a debit (negative) closing as assets here.
        # Banks with credit (positive) closing are picked up in bank_od_in_bank_accounts().
        cash = -self._sum_closing("Cash-in-hand")
        bank_asset = sum(-l.closing for l in self._ledgers_for("Bank Accounts")
                         if l.closing < 0)
        return cash + bank_asset

    def bank_od_in_bank_accounts(self) -> float:
        """Bank Accounts ledgers with credit (positive) closing = additional OD / liability."""
        return sum(l.closing for l in self._ledgers_for("Bank Accounts") if l.closing > 0)

    def duties_and_taxes_asset(self) -> float:
        """Net GST/TDS refund receivable (when D&T net is debit/negative)."""
        net = self.duties_and_taxes_net()
        return -net if net < 0 else 0.0

    def other_current_assets(self) -> float:
        oca       = -self._sum_closing("Current Assets")
        branch_dr = max(0, -self._sum_closing("Branch / Divisions"))
        dt_asset  = self.duties_and_taxes_asset()
        return oca + branch_dr + dt_asset

    def deferred_tax_asset(self) -> float:
        dtl = next((l.closing for l in self.ledgers
                    if "deferred tax" in l.name.lower()), 0.0)
        return max(0, -dtl)  # if DTL < 0, it's actually a DTA

    def total_noncurrent_assets(self) -> float:
        return (self.net_fixed_assets()
                + self.non_current_investments()
                + self.long_term_loans_advances()
                + self.other_noncurrent_assets()
                + self.deferred_tax_asset())

    def total_current_assets(self) -> float:
        return (self.closing_stock()
                + self.trade_receivables()
                + self.cash_and_bank()
                + self.other_current_assets())

    def total_assets(self) -> float:
        return self.total_noncurrent_assets() + self.total_current_assets()

    def total_equity(self) -> float:
        return self.share_capital() + self.reserves_surplus()

    def deferred_tax_liability(self) -> float:
        dtl = next((l.closing for l in self.ledgers
                    if "deferred tax" in l.name.lower()), 0.0)
        return max(0, dtl)

    def total_noncurrent_liab(self) -> float:
        return self.long_term_borrowings() + self.deferred_tax_liability()

    def total_current_liab(self) -> float:
        return (self.short_term_borrowings()
                + self.bank_od_in_bank_accounts()   # OD banks under Bank Accounts group
                + self.trade_payables()
                + max(0, self.duties_and_taxes_net())
                + self.other_current_liabilities()
                + self.short_term_provisions())

    def total_equity_liabilities(self) -> float:
        return self.total_equity() + self.total_noncurrent_liab() + self.total_current_liab()


# ─── Validation ───────────────────────────────────────────────────────────────

ERROR   = "ERROR"
WARNING = "WARNING"
INFO    = "INFO"

@dataclass
class Check:
    severity: str     # ERROR | WARNING | INFO
    category: str     # Schema | Balance | P&L | Data Quality | Classification
    message: str
    detail: str = ""

class ValidationResult:
    def __init__(self) -> None:
        self.checks: list[Check] = []

    def error(self, cat: str, msg: str, detail: str = "") -> None:
        self.checks.append(Check(ERROR, cat, msg, detail))

    def warning(self, cat: str, msg: str, detail: str = "") -> None:
        self.checks.append(Check(WARNING, cat, msg, detail))

    def info(self, cat: str, msg: str, detail: str = "") -> None:
        self.checks.append(Check(INFO, cat, msg, detail))

    @property
    def has_errors(self) -> bool:
        return any(c.severity == ERROR for c in self.checks)

    @property
    def errors(self) -> list[Check]:
        return [c for c in self.checks if c.severity == ERROR]

    @property
    def warnings(self) -> list[Check]:
        return [c for c in self.checks if c.severity == WARNING]

    def summary(self) -> str:
        e = len(self.errors)
        w = len(self.warnings)
        parts = []
        if e: parts.append(f"{e} error{'s' if e > 1 else ''}")
        if w: parts.append(f"{w} warning{'s' if w > 1 else ''}")
        return ", ".join(parts) if parts else "✓ All checks passed"


# ── Amount parser ─────────────────────────────────────────────────────────────

_INFER_KEYWORDS: list[tuple[list[str], str]] = [
    # P&L — Revenue
    (["sales account", "sales accounts", "revenue", "turnover"], "Sales Accounts"),
    (["direct income", "direct incomes"],                         "Direct Incomes"),
    (["indirect income", "indirect incomes", "other income"],     "Indirect Incomes"),
    # P&L — Expenses
    (["purchase account", "purchase accounts"],                   "Purchase Accounts"),
    (["direct expense", "direct expenses"],                       "Direct Expenses"),
    (["indirect expense", "indirect expenses"],                   "Indirect Expenses"),
    (["expense", "expenditure", "cost", "consumption"],           "Indirect Expenses"),
    # BS — Assets
    (["fixed asset", "tangible asset", "intangible asset",
      "capital work", "cwip"],                                     "Fixed Assets"),
    (["investment"],                                               "Investments"),
    (["stock", "inventory", "inventories"],                       "Stock-in-hand"),
    (["sundry debtor", "trade receivable", "debtor"],             "Sundry Debtors"),
    (["cash-in-hand", "cash in hand", "petty cash"],              "Cash-in-hand"),
    (["bank account", "bank"],                                    "Bank Accounts"),
    (["loan and advance", "loans and advance", "advance"],        "Loans & Advances (Asset)"),
    (["deposit"],                                                  "Deposits (Asset)"),
    (["current asset"],                                           "Current Assets"),
    (["misc. expense", "deferred expense", "preliminary"],        "Misc. Expenses (ASSET)"),
    # BS — Equity
    (["share capital", "equity capital", "paid-up capital",
      "capital account"],                                          "Capital Account"),
    (["reserve", "surplus"],                                      "Reserves & Surplus"),
    # BS — Liabilities
    (["secured loan", "term loan", "vehicle loan"],               "Secured Loans"),
    (["unsecured loan"],                                           "Unsecured Loans"),
    (["bank od", "overdraft", "cc limit", "cash credit",
      "working capital loan"],                                     "Bank OD A/c"),
    (["sundry creditor", "trade payable", "creditor"],            "Sundry Creditors"),
    (["provision"],                                                "Provisions"),
    (["duties", "taxes payable", "gst", "tds payable"],          "Duties & Taxes"),
    (["current liabilit", "other liabilit"],                      "Current Liabilities"),
    (["branch", "division"],                                      "Branch / Divisions"),
    (["suspense"],                                                 "Suspense A/c"),
]

def infer_standard_group(primary: str, parents: set[str]) -> str | None:
    """Heuristically map an unknown primary group to the nearest BS_MAP / PNL_MAP key.
    Returns None if no confident match is found.
    """
    text = f"{primary} {' '.join(parents)}".lower()
    for keywords, target in _INFER_KEYWORDS:
        if any(kw in text for kw in keywords):
            return target
    return None


def safe_float(value: Any, default: float = 0.0) -> float:
    """Parse a Tally TEXT amount field robustly.

    Handles: None, empty string, plain floats, Indian comma format
    (1,23,456.78), Cr/Dr suffixes, and unexpected non-numeric content.

    Tally sign convention (stored in the DB):
      negative → debit balance   positive → credit balance
    The Cr/Dr suffix in some older exports reverses this per usual
    double-entry convention, so we handle both.
    """
    if value is None:
        return default
    s = str(value).strip()
    if not s:
        return default
    # Strip Indian-format commas: "1,23,456.78" → "123456.78"
    s_clean = s.replace(",", "")
    low = s_clean.lower()
    # Handle "123456.78 Cr" / "123456.78 Dr" suffix (some Tally versions)
    if low.endswith("cr"):
        try:
            return float(low[:-2].strip())   # Cr in Tally = credit = positive
        except ValueError:
            return default
    if low.endswith("dr"):
        try:
            return -float(low[:-2].strip())  # Dr in Tally = debit = negative
        except ValueError:
            return default
    try:
        return float(s_clean)
    except (ValueError, TypeError):
        return default


# ── Schema validator ──────────────────────────────────────────────────────────

_REQUIRED_TABLES = {"mst_ledger", "mst_group", "_export_info"}
_OPTIONAL_TABLES = {"trn_accounting", "trn_voucher"}

_REQUIRED_LEDGER_COLS = {"name", "parent", "closing_balance"}
_REQUIRED_GROUP_COLS  = {"name", "primary_group"}


def validate_schema(con: sqlite3.Connection) -> ValidationResult:
    """Check that the database has the structure we expect."""
    vr = ValidationResult()

    existing = {r[0] for r in con.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    )}

    for t in _REQUIRED_TABLES:
        if t not in existing:
            vr.error("Schema", f"Required table missing: '{t}'",
                     "This may not be a valid Tally SQLite export from this app.")

    for t in _OPTIONAL_TABLES:
        if t not in existing:
            vr.warning("Schema", f"Optional table absent: '{t}'",
                       "Some cross-checks will be skipped.")

    if "mst_ledger" in existing:
        cols = {r[1] for r in con.execute("PRAGMA table_info(mst_ledger)")}
        for c in _REQUIRED_LEDGER_COLS:
            if c not in cols:
                vr.error("Schema", f"mst_ledger missing column '{c}'")
        if "opening_balance" not in cols:
            vr.warning("Schema", "mst_ledger has no 'opening_balance' column",
                       "Opening stock and prior-year P&L carry-forward will be zero.")

    if "mst_group" in existing:
        cols = {r[1] for r in con.execute("PRAGMA table_info(mst_group)")}
        for c in _REQUIRED_GROUP_COLS:
            if c not in cols:
                vr.error("Schema", f"mst_group missing column '{c}'")

    if "_export_info" in existing:
        cols = {r[1] for r in con.execute("PRAGMA table_info(_export_info)")}
        if "name" not in cols or "value" not in cols:
            vr.warning("Schema", "_export_info table has unexpected structure",
                       "Company name and period may not load correctly.")

    return vr


# ── Financial-data validator ──────────────────────────────────────────────────

def validate_financial_data(fd: "FinancialData") -> ValidationResult:
    """Business-logic checks on the loaded financial data."""
    vr = ValidationResult()

    # ── Balance sheet equation ────────────────────────────────────────────────
    diff = fd.total_assets() - fd.total_equity_liabilities()
    if abs(diff) < 1:
        vr.info("Balance Sheet", "Balance sheet balances to zero. ✓")
    elif abs(diff) < 50_000:
        vr.warning("Balance Sheet",
                   f"Small imbalance: ₹{diff:,.2f}",
                   "Likely a rounding difference in stock or depreciation entries.")
    else:
        vr.error("Balance Sheet",
                 f"Balance sheet does not balance — difference: ₹{diff:,.0f}",
                 "Check for ledgers with unusual group assignments or missing classifications.")

    # ── P&L reconciliation ────────────────────────────────────────────────────
    cy_profit    = fd.pnl_balance - fd.pnl_opening   # current year as Tally computed it
    our_pbt      = fd.profit_before_tax()
    pnl_diff     = abs(our_pbt - cy_profit)
    if pnl_diff < 1:
        vr.info("P&L", f"P&L reconciles exactly with Tally (₹{our_pbt:,.0f}). ✓")
    elif pnl_diff < 50_000:
        vr.warning("P&L",
                   f"Minor P&L gap: computed ₹{our_pbt:,.0f} vs Tally ₹{cy_profit:,.0f} "
                   f"(diff ₹{pnl_diff:,.0f})",
                   "May be caused by rounding in opening/closing stock or minor ledger mismatches.")
    else:
        vr.error("P&L",
                 f"P&L does not reconcile — computed ₹{our_pbt:,.0f}, "
                 f"Tally ₹{cy_profit:,.0f}, diff ₹{pnl_diff:,.0f}",
                 "Some ledgers may be under a group not mapped in BS_MAP or PNL_MAP. "
                 "See 'Unclassified' section below.")

    # ── Unclassified primary groups ───────────────────────────────────────────
    all_pgs = {l.primary_group for l in fd.ledgers if l.primary_group}
    for pg in sorted(all_pgs):
        if pg not in BS_MAP and pg not in PNL_MAP:
            net = sum(l.closing for l in fd.ledgers if l.primary_group == pg)
            if abs(net) > 1_000:
                vr.warning("Classification",
                           f"Unclassified primary group '{pg}': net ₹{net:,.0f}",
                           "Add this group to BS_MAP or PNL_MAP at the top of the script "
                           "to include it in statements.")

    # ── Stock checks ──────────────────────────────────────────────────────────
    if fd.opening_stock() == 0 and fd.closing_stock() > 0 and fd.pnl_balance != 0:
        vr.warning("Data Quality",
                   "Opening stock is zero but closing stock is non-zero",
                   "If this is not a new business, verify the opening stock ledger.")

    rev = fd.revenue_from_ops()
    cs  = fd.closing_stock()
    if rev > 0 and cs > rev * 0.6:
        vr.warning("Data Quality",
                   f"Closing stock ₹{cs:,.0f} is >60% of revenue ₹{rev:,.0f}",
                   "Verify the closing stock value entered is correct.")

    # ── Cash & bank ───────────────────────────────────────────────────────────
    if fd.cash_and_bank() < -10_000:
        vr.warning("Data Quality",
                   f"Net cash & bank is negative (₹{fd.cash_and_bank():,.0f})",
                   "Some bank accounts may have credit balances (OD) that are classified "
                   "under 'Bank Accounts' rather than 'Bank OD A/c'. Review Note 13.")

    # ── Receivables vs revenue ────────────────────────────────────────────────
    tr = fd.trade_receivables()
    if rev > 0 and tr > rev:
        vr.warning("Data Quality",
                   f"Trade receivables ₹{tr:,.0f} exceed annual revenue ₹{rev:,.0f}",
                   "Debtor days > 365. Check for stale/uncollected debtors or "
                   "misclassified items under Sundry Debtors.")

    # ── Capital & equity ──────────────────────────────────────────────────────
    if fd.share_capital() < 0:
        vr.error("Data Quality",
                 f"Share Capital is negative (₹{fd.share_capital():,.0f})",
                 "Capital Account ledgers have a net debit balance. "
                 "Check Capital Account entries in Tally.")

    if fd.reserves_surplus() < -100_000:
        vr.warning("Data Quality",
                   f"Reserves & Surplus is significantly negative (₹{fd.reserves_surplus():,.0f})",
                   "Accumulated losses exceed paid-up capital. May be technically insolvent.")

    # ── Borrowings sign check ─────────────────────────────────────────────────
    if fd.long_term_borrowings() < 0:
        vr.warning("Data Quality",
                   f"Long-term borrowings net is negative (₹{fd.long_term_borrowings():,.0f})",
                   "Secured/Unsecured Loan ledgers have a net debit balance. "
                   "Check if loan repayments exceeded drawdowns or if ledgers are miscoded.")

    # ── Period check ──────────────────────────────────────────────────────────
    try:
        pf = date.fromisoformat(fd.period_from)
        pt = date.fromisoformat(fd.period_to)
        days = (pt - pf).days
        if days < 300 or days > 400:
            vr.warning("Data Quality",
                       f"Period is {days} days ({fd.period_from} → {fd.period_to})",
                       "Expected a 12-month financial year (365 days). "
                       "Projections assume a full year as the base period.")
    except (ValueError, TypeError):
        vr.warning("Schema", "Could not parse period dates from _export_info.")

    # ── Significant ledgers with no BS/PL classification ─────────────────────
    unclass = [l for l in fd.ledgers
               if l.primary_group not in BS_MAP
               and l.primary_group not in PNL_MAP
               and abs(l.closing) > 50_000]
    if unclass:
        examples = "; ".join(f"{l.name} (₹{l.closing:,.0f})" for l in unclass[:4])
        vr.warning("Classification",
                   f"{len(unclass)} ledger(s) with material balances are unclassified",
                   f"Examples: {examples}")

    # ── Per-ledger sign/stale checks (from BalanceSheetCleanlinessAnalytics) ──
    asset_groups = {pg for pg, (side, _, _) in BS_MAP.items() if side in ("noncurrent_asset", "current_asset")}
    liab_groups  = {pg for pg, (side, _, _) in BS_MAP.items() if side in ("equity", "noncurrent_liab", "current_liab")}

    sign_flips, natural_breaches, stale = [], [], []
    for l in fd.ledgers:
        # Sign flip: opening and closing have opposite signs
        if abs(l.opening) > 1 and abs(l.closing) > 1 and l.opening * l.closing < 0:
            sign_flips.append(l.name)

        # Natural sign breach: asset with credit balance or liability with debit balance
        if l.primary_group in asset_groups and l.closing > 1_000:
            natural_breaches.append(f"{l.name} (asset with credit balance ₹{l.closing:,.0f})")
        if l.primary_group in liab_groups and l.closing < -1_000:
            natural_breaches.append(f"{l.name} (liability with debit balance ₹{l.closing:,.0f})")

        # Stale: large balance but opening == closing (no movement)
        if abs(l.closing) >= 100_000 and abs(l.opening - l.closing) < 1 and l.opening != 0:
            stale.append(f"{l.name} (₹{l.closing:,.0f})")

    if sign_flips:
        examples = "; ".join(sign_flips[:3])
        vr.warning("Data Quality",
                   f"{len(sign_flips)} ledger(s) have sign flips (opening/closing opposite signs)",
                   f"Examples: {examples}")
    if natural_breaches:
        examples = "; ".join(natural_breaches[:3])
        vr.warning("Data Quality",
                   f"{len(natural_breaches)} ledger(s) breach their natural sign",
                   f"Examples: {examples}")
    if stale:
        examples = "; ".join(stale[:3])
        vr.info("Data Quality",
                f"{len(stale)} ledger(s) have large stale balances (no movement in period)",
                f"Examples: {examples}")

    return vr


# ─── Database loader ──────────────────────────────────────────────────────────

def load_from_sqlite(db_path: str,
                     opening_stock_override: float | None = None,
                     closing_stock_override: float | None = None,
                     schema_vr: ValidationResult | None = None,
                     reclassify_map: dict[str, str] | None = None) -> FinancialData:
    """Load financial data from a Tally SQLite export.

    Raises RuntimeError if schema validation finds blocking errors.
    Non-fatal warnings are accumulated into schema_vr (passed in or a new one).
    """
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row

    # Schema validation — fail fast on missing tables / columns
    sv = validate_schema(con)
    if schema_vr is not None:
        schema_vr.checks.extend(sv.checks)
    if sv.has_errors:
        con.close()
        msgs = "; ".join(c.message for c in sv.errors)
        raise RuntimeError(f"Database schema errors: {msgs}")

    # Company / period (_export_info uses name/value columns)
    try:
        info = {r["name"]: r["value"]
                for r in con.execute("SELECT name, value FROM _export_info")}
    except Exception:
        info = {}
    company     = info.get("company_name", "Company").replace("(from", "").replace(")", "").strip()
    period_from = info.get("period_from", "")
    period_to   = info.get("period_to",   "")

    # Build period label e.g. "31st March 2026"
    def _ordinal(dt_str: str) -> str:
        try:
            d = date.fromisoformat(dt_str)
            day = d.day
            suffix = {1:"st",2:"nd",3:"rd"}.get(day if day < 20 else day % 10, "th")
            return f"{day}{suffix} {d.strftime('%B %Y')}"
        except Exception:
            return dt_str
    period_label = _ordinal(period_to)

    # Detect whether opening_balance column exists (older exports may omit it)
    ledger_cols = {r[1] for r in con.execute("PRAGMA table_info(mst_ledger)")}
    has_opening = "opening_balance" in ledger_cols

    # Profit & Loss A/c (special ledger, no parent group)
    # closing_balance = cumulative (prior year opening + current year profit)
    # opening_balance = prior year retained earnings not yet transferred to Reserves
    pnl_balance = 0.0
    pnl_opening = 0.0
    try:
        if has_opening:
            pnl_row = con.execute(
                "SELECT opening_balance, closing_balance FROM mst_ledger "
                "WHERE name = 'Profit & Loss A/c'"
            ).fetchone()
            if pnl_row:
                pnl_balance = safe_float(pnl_row["closing_balance"])
                pnl_opening = safe_float(pnl_row["opening_balance"])
        else:
            pnl_row = con.execute(
                "SELECT closing_balance FROM mst_ledger WHERE name = 'Profit & Loss A/c'"
            ).fetchone()
            if pnl_row:
                pnl_balance = safe_float(pnl_row["closing_balance"])
    except Exception:
        pass   # pnl_balance stays 0

    # Balance-sheet ledgers from mst_ledger (closing balances = year-end positions)
    open_col = "l.opening_balance" if has_opening else "'0'"
    bs_rows = con.execute(f"""
        SELECT
            l.name,
            l.parent,
            g.primary_group,
            {open_col} AS opening_raw,
            l.closing_balance AS closing_raw,
            COALESCE(g.is_deemedpositive, '0') AS is_dp
        FROM mst_ledger l
        LEFT JOIN mst_group g ON g.name = l.parent
        WHERE g.primary_group IS NOT NULL
        ORDER BY g.primary_group, l.parent, l.name
    """).fetchall()

    # Build parent sets per primary for auto-inference
    _pg_parents: dict[str, set[str]] = {}
    for r in bs_rows:
        pg = r["primary_group"]
        if pg:
            _pg_parents.setdefault(pg, set()).add(str(r["parent"] or ""))

    effective_remap = dict(reclassify_map or {})

    ledgers = []
    for r in bs_rows:
        pg = r["primary_group"]
        # Apply user reclassification first
        if pg in effective_remap:
            pg = effective_remap[pg]
        # Auto-infer for groups still not in any map
        elif pg not in BS_MAP and pg not in PNL_MAP:
            inferred = infer_standard_group(pg, _pg_parents.get(r["primary_group"], set()))
            if inferred:
                pg = inferred
        ledgers.append(LedgerRow(
            name=r["name"],
            parent=r["parent"],
            primary_group=pg,
            opening=safe_float(r["opening_raw"]),
            closing=safe_float(r["closing_raw"]),
            is_deemedpositive=str(r["is_dp"]) == "1",
        ))

    # P&L activity from trn_accounting (kept for future use / cross-checks).
    # The primary P&L computation uses mst_ledger closing_balance instead
    # because trn_accounting includes journal/inter-branch entries that Tally
    # excludes from its own P&L report.
    pnl_groups = (
        "'Sales Accounts','Direct Incomes','Indirect Incomes',"
        "'Purchase Accounts','Direct Expenses','Indirect Expenses'"
    )
    try:
        pnl_txn_rows = con.execute(f"""
            SELECT
                g.primary_group,
                l.parent,
                SUM(CAST(a.amount AS REAL)) AS total
            FROM trn_accounting a
            JOIN mst_ledger l ON l.name = a.ledger
            LEFT JOIN mst_group g ON g.name = l.parent
            WHERE g.primary_group IN ({pnl_groups})
            GROUP BY g.primary_group, l.parent
            ORDER BY g.primary_group, l.parent
        """).fetchall()
        pnl_rows = [
            PnlRow(
                primary_group=r["primary_group"],
                parent=r["parent"],
                total=float(r["total"] or 0),
            )
            for r in pnl_txn_rows
        ]
    except Exception:
        pnl_rows = []   # trn_accounting absent or malformed — non-fatal

    con.close()
    return FinancialData(
        company=company,
        period_from=period_from,
        period_to=period_to,
        period_label=period_label,
        ledgers=ledgers,
        pnl_rows=pnl_rows,
        pnl_balance=pnl_balance,
        pnl_opening=pnl_opening,
        opening_stock_override=opening_stock_override,
        closing_stock_override=closing_stock_override,
    )


# ─── Excel generation ─────────────────────────────────────────────────────────

# Colour palette
C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2F5496"
C_LIGHT_BLUE = "D6E4F0"
C_HEADER_BG  = "1F3864"
C_SUBHD_BG   = "BDD7EE"
C_TOTAL_BG   = "D9E1F2"
C_WHITE      = "FFFFFF"
C_BLACK      = "000000"
C_AMBER      = "FFF2CC"

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border(style: str = "thin") -> Border:
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, size=10, color=C_BLACK, name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

INR = '#,##0'          # Indian number format (openpyxl uses comma-separated)
INR2 = '#,##0.00'

def _fmt(ws, cell_ref: str, value: float | None, italic: bool = False) -> None:
    cell = ws[cell_ref]
    if value is not None:
        cell.value = value
    cell.number_format = INR
    if italic:
        cell.font = Font(name="Calibri", size=10, italic=True)


class ExcelWriter:
    def __init__(self, fd: FinancialData, proj: "ProjectionInputs | None" = None):
        self.fd = fd
        self.proj = proj
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # remove default sheet

    def save(self, path: str) -> None:
        self._write_bs()
        self._write_pnl()
        # Note sheets (referenced by hyperlinks in BS face)
        self._write_note_share_capital()
        self._write_note_reserves()
        self._write_note_lt_borrowings()
        self._write_note_st_borrowings()
        self._write_note_trade_payables()
        self._write_note_fixed_assets()
        self._write_note_inventories()
        self._write_note_trade_receivables()
        self._write_note_cash_bank()
        self._write_notes_index()
        if self.proj:
            pe = ProjectionEngine(self.fd, self.proj)
            self._write_proj_pnl(pe)
            self._write_proj_bs(pe)
            self._write_assumptions()
        vr = validate_financial_data(self.fd)
        self._write_validation(vr)
        self.wb.save(path)

    # ── Validation sheet ──────────────────────────────────────────────────────

    def _write_validation(self, vr: ValidationResult) -> None:
        ws = self.wb.create_sheet("Validation")
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 60

        # Header
        ws.merge_cells("A1:D1")
        ws["A1"].value = f"DATA VALIDATION REPORT — {self.fd.company} — {self.fd.period_label}"
        ws["A1"].font = _font(bold=True, size=12, color=C_WHITE)
        ws["A1"].fill = _fill(C_HEADER_BG)
        ws["A1"].alignment = _align("center")

        ws["A2"].value = vr.summary()
        ws["A2"].font = _font(bold=True, size=11,
                              color="CC0000" if vr.has_errors else "2E7D32")
        ws.merge_cells("A2:D2")
        ws["A2"].alignment = _align("center")

        # Column headers
        for col, lbl in [("A", "Severity"), ("B", "Category"),
                          ("C", "Message"), ("D", "Detail")]:
            ws[f"{col}3"].value = lbl
            ws[f"{col}3"].font = _font(bold=True, size=9, color=C_WHITE)
            ws[f"{col}3"].fill = _fill(C_MID_BLUE)
            ws[f"{col}3"].alignment = _align("center")
            ws[f"{col}3"].border = _border()

        _SEVERITY_COLORS = {ERROR: "FFCCCC", WARNING: C_AMBER, INFO: "E8F5E9"}
        _SEVERITY_FG     = {ERROR: "CC0000", WARNING: "7B5800", INFO: "1B5E20"}

        for i, chk in enumerate(vr.checks, start=4):
            bg = _SEVERITY_COLORS.get(chk.severity, C_WHITE)
            fg = _SEVERITY_FG.get(chk.severity, C_BLACK)
            for col, val in [("A", chk.severity), ("B", chk.category),
                              ("C", chk.message),  ("D", chk.detail)]:
                cell = ws[f"{col}{i}"]
                cell.value = val
                cell.fill = _fill(bg)
                cell.font = _font(size=9,
                                  bold=(col == "A"),
                                  color=fg if col == "A" else C_BLACK)
                cell.alignment = _align("left", wrap=True)
                cell.border = _border()
            ws.row_dimensions[i].height = 28

        ws.freeze_panes = "A4"

    # ── Balance Sheet ─────────────────────────────────────────────────────────

    def _write_bs(self) -> None:
        ws = self.wb.create_sheet("Balance Sheet")
        fd = self.fd
        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        r = 1
        def header(text, span=4, bg=C_HEADER_BG, fg=C_WHITE, sz=12, bold=True):
            nonlocal r
            ws.merge_cells(f"A{r}:D{r}")
            cell = ws[f"A{r}"]
            cell.value = text
            cell.font = _font(bold=bold, size=sz, color=fg)
            cell.fill = _fill(bg)
            cell.alignment = _align("center")
            r += 1

        def subheader(text, bg=C_SUBHD_BG):
            nonlocal r
            ws.merge_cells(f"A{r}:D{r}")
            cell = ws[f"A{r}"]
            cell.value = text
            cell.font = _font(bold=True, size=10)
            cell.fill = _fill(bg)
            cell.alignment = _align("left")
            r += 1

        def col_header():
            nonlocal r
            ws[f"A{r}"].value = "Particulars"
            ws[f"B{r}"].value = "Note"
            ws[f"C{r}"].value = f"As at {fd.period_label}"
            ws[f"D{r}"].value = f"Previous Year"
            for col in "ABCD":
                c = ws[f"{col}{r}"]
                c.font = _font(bold=True, size=9, color=C_WHITE)
                c.fill = _fill(C_MID_BLUE)
                c.alignment = _align("center")
                c.border = _border()
            r += 1

        def row(label, amount, note=None, indent=0, bold=False, total=False, fmt=INR):
            nonlocal r
            prefix = "  " * indent
            ws[f"A{r}"].value = prefix + label
            ws[f"A{r}"].font = _font(bold=bold or total, size=10)
            ws[f"A{r}"].alignment = _align("left")
            if note:
                note_sheet_name = self._note_sheet_name(int(note))
                ws[f"B{r}"].value = note
                ws[f"B{r}"].hyperlink = f"#'{note_sheet_name}'!A1"
                ws[f"B{r}"].font = Font(name="Calibri", size=9, color="1D4ED8", underline="single")
                ws[f"B{r}"].alignment = _align("center")
            if amount is not None:
                ws[f"C{r}"].value = amount
                ws[f"C{r}"].number_format = fmt
                ws[f"C{r}"].font = _font(bold=bold or total, size=10)
                ws[f"C{r}"].alignment = _align("right")
            if total:
                for col in "AC":
                    ws[f"{col}{r}"].fill = _fill(C_TOTAL_BG)
                    ws[f"{col}{r}"].border = _border()
            r += 1
            return r - 1  # return row number for formula references

        def spacer():
            nonlocal r
            r += 1

        header(fd.company.upper(), sz=14)
        header("BALANCE SHEET", sz=12)
        header(f"As at {fd.period_label}", sz=10, bg=C_MID_BLUE)
        header("(Amount in ₹)", sz=9, bg=C_LIGHT_BLUE, fg=C_BLACK, bold=False)
        col_header()

        # ── EQUITY & LIABILITIES ──────────────────────────────────────────────
        subheader("I.  SHAREHOLDERS' FUNDS")
        sc_r  = row("  a)  Share Capital",           fd.share_capital(),       note="1", indent=0)
        res_r = row("  b)  Reserves & Surplus",      fd.reserves_surplus(),    note="2", indent=0)
        row("Total Shareholders' Funds",
            None, bold=True, total=True)
        ws[f"C{r-1}"].value = f"=C{sc_r}+C{res_r}"
        ws[f"C{r-1}"].number_format = INR
        spacer()

        subheader("II.  NON-CURRENT LIABILITIES")
        ltb_r = row("  a)  Long-Term Borrowings",    fd.long_term_borrowings(), note="3", indent=0)
        dtl_r = row("  b)  Deferred Tax Liability",  fd.deferred_tax_liability(), indent=0)
        row("Total Non-Current Liabilities",
            None, bold=True, total=True)
        ws[f"C{r-1}"].value = f"=C{ltb_r}+C{dtl_r}"
        ws[f"C{r-1}"].number_format = INR
        spacer()

        subheader("III.  CURRENT LIABILITIES")
        stb_r  = row("  a)  Short-Term Borrowings",  fd.short_term_borrowings(), note="4", indent=0)
        tp_r   = row("  b)  Trade Payables",         fd.trade_payables(),         note="5", indent=0)
        dt_r   = row("  c)  Duties & Taxes (Net)",   max(0, fd.duties_and_taxes_net()), indent=0)
        ocl_r  = row("  d)  Other Current Liabilities", fd.other_current_liabilities(), note="6", indent=0)
        prov_r = row("  e)  Short-Term Provisions",  fd.short_term_provisions(),  note="7", indent=0)
        row("Total Current Liabilities",
            None, bold=True, total=True)
        ws[f"C{r-1}"].value = f"=C{stb_r}+C{tp_r}+C{dt_r}+C{ocl_r}+C{prov_r}"
        ws[f"C{r-1}"].number_format = INR
        total_cl_r = r - 1
        spacer()

        # Grand total E&L
        row("TOTAL EQUITY & LIABILITIES", None, bold=True)
        ws[f"C{r-1}"].value = fd.total_equity_liabilities()
        ws[f"C{r-1}"].number_format = INR
        ws[f"C{r-1}"].font = _font(bold=True, size=11)
        ws[f"C{r-1}"].fill = _fill(C_HEADER_BG)
        ws[f"C{r-1}"].font = Font(bold=True, size=11, color=C_WHITE, name="Calibri")
        ws[f"A{r-1}"].fill = _fill(C_HEADER_BG)
        ws[f"A{r-1}"].font = Font(bold=True, size=11, color=C_WHITE, name="Calibri")
        total_el_row = r - 1
        spacer(); spacer()

        # ── ASSETS ───────────────────────────────────────────────────────────
        subheader("I.  NON-CURRENT ASSETS")
        fa_r   = row("  a)  Fixed Assets (Net Block)",     fd.net_fixed_assets(),          note="8", indent=0)
        inv_r  = row("  b)  Non-Current Investments",      fd.non_current_investments(),   note="9", indent=0)
        lla_r  = row("  c)  Long-Term Loans & Advances",   fd.long_term_loans_advances(),  note="10", indent=0)
        dta_r  = row("  d)  Deferred Tax Asset",           fd.deferred_tax_asset(),        indent=0)
        ona_r  = row("  e)  Other Non-Current Assets",     fd.other_noncurrent_assets(),   indent=0)
        row("Total Non-Current Assets",
            None, bold=True, total=True)
        ws[f"C{r-1}"].value = f"=C{fa_r}+C{inv_r}+C{lla_r}+C{dta_r}+C{ona_r}"
        ws[f"C{r-1}"].number_format = INR
        spacer()

        subheader("II.  CURRENT ASSETS")
        stk_r  = row("  a)  Inventories (Closing Stock)",  fd.closing_stock(),             note="11", indent=0)
        tr_r   = row("  b)  Trade Receivables",            fd.trade_receivables(),         note="12", indent=0)
        cb_r   = row("  c)  Cash & Cash Equivalents",      fd.cash_and_bank(),             note="13", indent=0)
        oca_r  = row("  d)  Other Current Assets",         fd.other_current_assets(),      note="14", indent=0)
        row("Total Current Assets",
            None, bold=True, total=True)
        ws[f"C{r-1}"].value = f"=C{stk_r}+C{tr_r}+C{cb_r}+C{oca_r}"
        ws[f"C{r-1}"].number_format = INR
        spacer()

        row("TOTAL ASSETS", None, bold=True)
        ws[f"C{r-1}"].value = fd.total_assets()
        ws[f"C{r-1}"].number_format = INR
        ws[f"C{r-1}"].font = Font(bold=True, size=11, color=C_WHITE, name="Calibri")
        ws[f"C{r-1}"].fill = _fill(C_HEADER_BG)
        ws[f"A{r-1}"].fill = _fill(C_HEADER_BG)
        ws[f"A{r-1}"].font = Font(bold=True, size=11, color=C_WHITE, name="Calibri")
        spacer(); spacer()

        # ── Difference check ─────────────────────────────────────────────────
        diff = fd.total_assets() - fd.total_equity_liabilities()
        row("Balance Sheet Difference (should be 0)", diff,
            bold=True if abs(diff) > 1 else False)
        if abs(diff) > 1:
            ws[f"C{r-1}"].fill = _fill("FF0000")
            ws[f"C{r-1}"].font = Font(bold=True, size=10, color=C_WHITE)
        spacer(); spacer()

        ws.freeze_panes = "A6"

    # ── Note sheet helpers ────────────────────────────────────────────────────

    def _note_sheet_name(self, note_num: int | str) -> str:
        """Return the sheet name for a given note number."""
        _NOTE_TITLES = {
            1: "N1 Share Capital",
            2: "N2 Reserves Surplus",
            3: "N3 LT Borrowings",
            4: "N4 ST Borrowings",
            5: "N5 Trade Payables",
            6: "N6 Other CL",
            7: "N7 Provisions",
            8: "N8 Fixed Assets",
            9: "N9 NC Investments",
            10: "N10 LT Loans",
            11: "N11 Inventories",
            12: "N12 Trade Receivables",
            13: "N13 Cash & Bank",
            14: "N14 Other CA",
        }
        return _NOTE_TITLES.get(int(note_num), f"Note {note_num}")

    def _start_note_sheet(self, note_num: int, title: str) -> tuple:
        """Create a note sheet, write header + back link, return (ws, next_row)."""
        sheet_name = self._note_sheet_name(note_num)
        ws = self.wb.create_sheet(sheet_name)
        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20

        # Title row
        ws.merge_cells("A1:C1")
        ws["A1"].value = f"Note {note_num}:  {title}"
        ws["A1"].font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
        ws["A1"].fill = _fill(C_HEADER_BG)
        ws["A1"].alignment = _align("center")

        # Back link
        ws.merge_cells("A2:C2")
        ws["A2"].value = "<- Back to Balance Sheet"
        ws["A2"].hyperlink = "#'Balance Sheet'!A1"
        ws["A2"].font = Font(name="Calibri", size=9, color="1D4ED8", underline="single")
        ws["A2"].alignment = _align("left")

        return ws, 3   # next row to write at

    def _note_data_row(self, ws, r: int, label: str, amount: float | None = None,
                       bold: bool = False, total: bool = False, indent: int = 0) -> int:
        prefix = "    " * indent
        ws[f"A{r}"].value = prefix + label
        ws[f"A{r}"].font = _font(bold=bold or total, size=9)
        if amount is not None:
            ws[f"B{r}"].value = amount
            ws[f"B{r}"].number_format = INR
            ws[f"B{r}"].font = _font(bold=bold or total, size=9)
            ws[f"B{r}"].alignment = _align("right")
        if total:
            ws[f"A{r}"].fill = _fill(C_TOTAL_BG)
            ws[f"B{r}"].fill = _fill(C_TOTAL_BG)
        return r + 1

    def _write_note_share_capital(self):
        ws, r = self._start_note_sheet(1, "Share Capital")
        fd = self.fd
        for l in fd._ledgers_for("Capital Account"):
            if "reserve" not in l.name.lower() and "profit" not in l.name.lower():
                r = self._note_data_row(ws, r, l.name, l.closing)
        r = self._note_data_row(ws, r, "Total Share Capital", fd.share_capital(), total=True)

    def _write_note_reserves(self):
        ws, r = self._start_note_sheet(2, "Reserves & Surplus")
        fd = self.fd
        for l in fd._ledgers_for("Capital Account"):
            if "reserve" in l.name.lower():
                r = self._note_data_row(ws, r, l.name, l.closing)
        for l in fd._ledgers_for("Reserves & Surplus"):
            r = self._note_data_row(ws, r, l.name, l.closing)
        r = self._note_data_row(ws, r, "Profit for the year (P&L A/c)", fd.pnl_balance)
        r = self._note_data_row(ws, r, "Total Reserves & Surplus", fd.reserves_surplus(), total=True)

    def _write_note_lt_borrowings(self):
        ws, r = self._start_note_sheet(3, "Long-Term Borrowings")
        fd = self.fd
        for pg in ("Secured Loans", "Unsecured Loans", "Loans (Liability)"):
            group_ledgers = fd._ledgers_for(pg)
            if not group_ledgers:
                continue
            r = self._note_data_row(ws, r, pg, bold=True)
            for l in group_ledgers:
                r = self._note_data_row(ws, r, l.name, l.closing, indent=1)
        r = self._note_data_row(ws, r, "Total Long-Term Borrowings", fd.long_term_borrowings(), total=True)

    def _write_note_st_borrowings(self):
        ws, r = self._start_note_sheet(4, "Short-Term Borrowings")
        fd = self.fd
        for l in fd._ledgers_for("Bank OD A/c"):
            r = self._note_data_row(ws, r, l.name, l.closing)
        # Bank accounts with credit balance = OD
        od_banks = [l for l in fd._ledgers_for("Bank Accounts") if l.closing > 0]
        if od_banks:
            r = self._note_data_row(ws, r, "Bank Accounts (credit/OD balance)", bold=True)
            for l in od_banks:
                r = self._note_data_row(ws, r, l.name, l.closing, indent=1)
        r = self._note_data_row(ws, r, "Total Short-Term Borrowings",
                                fd.short_term_borrowings() + fd.bank_od_in_bank_accounts(), total=True)

    def _write_note_trade_payables(self):
        ws, r = self._start_note_sheet(5, "Trade Payables")
        fd = self.fd
        by_parent: dict[str, float] = {}
        for l in fd._ledgers_for("Sundry Creditors"):
            by_parent[l.parent] = by_parent.get(l.parent, 0) + l.closing
        for parent, amt in sorted(by_parent.items()):
            r = self._note_data_row(ws, r, parent, amt if amt > 0 else None)
        r = self._note_data_row(ws, r, "Total Trade Payables", fd.trade_payables(), total=True)

    def _write_note_fixed_assets(self):
        ws, r = self._start_note_sheet(8, "Fixed Assets")
        fd = self.fd
        # Column headers
        for col, lbl in [("A", "Asset Category"), ("B", "Gross Open"), ("C", "Additions"),
                         ("D", "Disposals"), ("E", "Gross Close"), ("F", "Accum Depr"), ("G", "Net Block")]:
            ws.column_dimensions[col].width = 18
            ws[f"{col}{r}"].value = lbl
            ws[f"{col}{r}"].font = _font(bold=True, size=9, color=C_WHITE)
            ws[f"{col}{r}"].fill = _fill(C_MID_BLUE)
            ws[f"{col}{r}"].alignment = _align("center" if col != "A" else "left")
        r += 1
        for fa in fd.fixed_assets_schedule():
            ws[f"A{r}"].value = "  " + fa["name"]
            ws[f"A{r}"].font = _font(size=9)
            for col, key in [("B","gross_open"),("C","additions"),("D","disposals"),
                              ("E","gross_close"),("F","depr_close"),("G","net_close")]:
                ws[f"{col}{r}"].value = fa[key]
                ws[f"{col}{r}"].number_format = INR
                ws[f"{col}{r}"].font = _font(size=9)
                ws[f"{col}{r}"].alignment = _align("right")
            r += 1
        # Total row
        ws[f"A{r}"].value = "  TOTAL"
        ws[f"A{r}"].font = _font(bold=True, size=9)
        ws[f"G{r}"].value = fd.net_fixed_assets()
        ws[f"G{r}"].number_format = INR
        ws[f"G{r}"].font = _font(bold=True, size=9)
        ws[f"G{r}"].fill = _fill(C_TOTAL_BG)
        ws[f"A{r}"].fill = _fill(C_TOTAL_BG)

    def _write_note_inventories(self):
        ws, r = self._start_note_sheet(11, "Inventories")
        fd = self.fd
        r = self._note_data_row(ws, r, "Opening Stock (as per books / override)", fd.opening_stock())
        r = self._note_data_row(ws, r, "Closing Stock (as per books / override)", fd.closing_stock())
        r = self._note_data_row(ws, r, "Net Inventory on Balance Sheet", fd.closing_stock(), total=True)

    def _write_note_trade_receivables(self):
        ws, r = self._start_note_sheet(12, "Trade Receivables")
        fd = self.fd
        by_parent: dict[str, float] = {}
        for l in fd._ledgers_for("Sundry Debtors"):
            by_parent[l.parent] = by_parent.get(l.parent, 0) + (-l.closing)
        for parent, amt in sorted(by_parent.items()):
            r = self._note_data_row(ws, r, parent, amt if abs(amt) > 0 else None)
        r = self._note_data_row(ws, r, "Total Trade Receivables", fd.trade_receivables(), total=True)

    def _write_note_cash_bank(self):
        ws, r = self._start_note_sheet(13, "Cash & Cash Equivalents")
        fd = self.fd
        if fd._ledgers_for("Cash-in-hand"):
            r = self._note_data_row(ws, r, "Cash-in-Hand", bold=True)
            for l in fd._ledgers_for("Cash-in-hand"):
                r = self._note_data_row(ws, r, l.name, -l.closing, indent=1)
        asset_banks = [l for l in fd._ledgers_for("Bank Accounts") if l.closing < 0]
        if asset_banks:
            r = self._note_data_row(ws, r, "Bank Accounts (debit balance)", bold=True)
            for l in asset_banks:
                r = self._note_data_row(ws, r, l.name, -l.closing, indent=1)
        r = self._note_data_row(ws, r, "Total Cash & Cash Equivalents", fd.cash_and_bank(), total=True)

    def _write_notes_index(self):
        ws = self.wb.create_sheet("Notes Index")
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 20

        ws.merge_cells("A1:C1")
        ws["A1"].value = "NOTES TO FINANCIAL STATEMENTS"
        ws["A1"].font = _font(bold=True, size=11, color=C_WHITE)
        ws["A1"].fill = _fill(C_HEADER_BG)
        ws["A1"].alignment = _align("center")

        for col, lbl in [("A","Note No."),("B","Description"),("C","Amount (Rs.)")]:
            ws[f"{col}2"].value = lbl
            ws[f"{col}2"].font = _font(bold=True, size=9, color=C_WHITE)
            ws[f"{col}2"].fill = _fill(C_MID_BLUE)
            ws[f"{col}2"].alignment = _align("center")

        fd = self.fd
        NOTE_INDEX = [
            (1,  "Share Capital",                  fd.share_capital()),
            (2,  "Reserves & Surplus",             fd.reserves_surplus()),
            (3,  "Long-Term Borrowings",           fd.long_term_borrowings()),
            (4,  "Short-Term Borrowings",          fd.short_term_borrowings() + fd.bank_od_in_bank_accounts()),
            (5,  "Trade Payables",                 fd.trade_payables()),
            (8,  "Fixed Assets (Net Block)",       fd.net_fixed_assets()),
            (11, "Inventories",                    fd.closing_stock()),
            (12, "Trade Receivables",              fd.trade_receivables()),
            (13, "Cash & Cash Equivalents",        fd.cash_and_bank()),
        ]
        for i, (num, title, amount) in enumerate(NOTE_INDEX, start=3):
            sheet_name = self._note_sheet_name(num)
            ws[f"A{i}"].value = f"Note {num}"
            ws[f"A{i}"].hyperlink = f"#'{sheet_name}'!A1"
            ws[f"A{i}"].font = Font(name="Calibri", size=9, color="1D4ED8", underline="single")
            ws[f"A{i}"].alignment = _align("center")
            ws[f"B{i}"].value = title
            ws[f"B{i}"].font = _font(size=9)
            ws[f"C{i}"].value = amount
            ws[f"C{i}"].number_format = INR
            ws[f"C{i}"].font = _font(size=9)
            ws[f"C{i}"].alignment = _align("right")
            if i % 2 == 0:
                for col in "ABC":
                    ws[f"{col}{i}"].fill = _fill(C_LIGHT_BLUE)

        ws.freeze_panes = "A3"

    # ── P&L Statement ─────────────────────────────────────────────────────────

    def _write_pnl(self) -> None:
        ws = self.wb.create_sheet("P&L Statement")
        fd = self.fd
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 18

        r = 1

        def header(text, bg=C_HEADER_BG, fg=C_WHITE, sz=12, bold=True):
            nonlocal r
            ws.merge_cells(f"A{r}:C{r}")
            c = ws[f"A{r}"]
            c.value = text
            c.font = _font(bold=bold, size=sz, color=fg)
            c.fill = _fill(bg)
            c.alignment = _align("center")
            r += 1

        def row(label, amount=None, bold=False, total=False, indent=0, italic=False, bg=None):
            nonlocal r
            ws[f"A{r}"].value = "  " * indent + label
            ws[f"A{r}"].font = _font(bold=bold or total, size=10, name="Calibri")
            if italic:
                ws[f"A{r}"].font = Font(italic=True, size=10, name="Calibri")
            if amount is not None:
                ws[f"C{r}"].value = amount
                ws[f"C{r}"].number_format = INR
                ws[f"C{r}"].font = _font(bold=bold or total, size=10)
                ws[f"C{r}"].alignment = _align("right")
            if total or bg:
                col_bg = bg or C_TOTAL_BG
                ws[f"A{r}"].fill = _fill(col_bg)
                ws[f"C{r}"].fill = _fill(col_bg)
            r += 1
            return r - 1

        def spacer():
            nonlocal r; r += 1

        header(fd.company.upper(), sz=13)
        header("STATEMENT OF PROFIT & LOSS", sz=11)
        header(f"For the Year Ended {fd.period_label}", sz=10, bg=C_MID_BLUE)
        header("(Amount in ₹)", sz=9, bg=C_LIGHT_BLUE, fg=C_BLACK, bold=False)

        # Column labels
        ws[f"A{r}"].value = "Particulars"
        ws[f"B{r}"].value = "Note"
        ws[f"C{r}"].value = "Current Year"
        for col in "ABC":
            c = ws[f"{col}{r}"]
            c.font = _font(bold=True, size=9, color=C_WHITE)
            c.fill = _fill(C_MID_BLUE)
            c.alignment = _align("center")
            c.border = _border()
        r += 1

        # ── Revenue ──────────────────────────────────────────────────────────
        rev_r  = row("I.   Revenue from Operations",  fd.revenue_from_ops(), bold=True)
        oth_r  = row("II.  Other Income",              fd.other_income(),    bold=True)
        tot_r  = row("III. Total Revenue (I + II)",    None, bold=True, total=True)
        ws[f"C{tot_r}"].value  = f"=C{rev_r}+C{oth_r}"
        ws[f"C{tot_r}"].number_format = INR
        spacer()

        # ── Expenses ─────────────────────────────────────────────────────────
        row("IV.  Expenses", bold=True)
        pur_r  = row("     a)  Cost of Materials / Purchases", fd.purchases(), indent=1)
        chg_r  = row("     b)  Changes in Inventories",
                      fd.opening_stock() - fd.closing_stock(), indent=1,
                      italic=True)
        row("          (Opening Stock - Closing Stock)", italic=True, indent=2)
        emp_r  = row("     c)  Employee Benefits Expense",  fd.employee_costs(), indent=1)
        fin_r  = row("     d)  Finance Costs",              fd.finance_costs(), indent=1)
        dep_r  = row("     e)  Depreciation & Amortisation",fd.depreciation_from_fa(), indent=1)
        oth2_r = row("     f)  Other Expenses",             fd.other_indirect_expenses() + fd.direct_expenses(), indent=1)

        # Show breakdown of Other Expenses
        row("          Direct Expenses", fd.direct_expenses(), indent=3, italic=True)
        row("          Indirect Expenses (excl. Finance & Employee)", fd.other_indirect_expenses(), indent=3, italic=True)

        tot_exp_r = row("     Total Expenses", None, bold=True, total=True)
        ws[f"C{tot_exp_r}"].value = (
            f"=C{pur_r}+C{chg_r}+C{emp_r}+C{fin_r}+C{dep_r}+C{oth2_r}"
        )
        ws[f"C{tot_exp_r}"].number_format = INR
        spacer()

        # ── Profit lines ─────────────────────────────────────────────────────
        pbt_r = row("V.   Profit Before Tax (III − Expenses)", None, bold=True, total=True)
        ws[f"C{pbt_r}"].value = f"=C{tot_r}-C{tot_exp_r}"
        ws[f"C{pbt_r}"].number_format = INR

        tax_r = row("VI.  Tax Expense (Deferred Tax Provision)", fd.tax_expense())
        pat_r = row("VII. Profit After Tax", None, bold=True, bg=C_HEADER_BG)
        ws[f"C{pat_r}"].value = f"=C{pbt_r}-C{tax_r}"
        ws[f"C{pat_r}"].number_format = INR
        ws[f"C{pat_r}"].font = Font(bold=True, size=11, color=C_WHITE, name="Calibri")
        ws[f"A{pat_r}"].font = Font(bold=True, size=11, color=C_WHITE, name="Calibri")
        spacer(); spacer()

        # ── Key ratios ───────────────────────────────────────────────────────
        ws.merge_cells(f"A{r}:C{r}")
        ws[f"A{r}"].value = "KEY FINANCIAL RATIOS"
        ws[f"A{r}"].font = _font(bold=True, size=10, color=C_WHITE)
        ws[f"A{r}"].fill = _fill(C_MID_BLUE)
        r += 1

        rev = fd.revenue_from_ops()
        gp  = rev - fd.purchases() + (fd.closing_stock() - fd.opening_stock())
        ebitda = fd.profit_before_tax() + fd.finance_costs() + fd.depreciation_from_fa()
        ratios = [
            ("Gross Profit", gp, f"{gp/rev*100:.1f}% of Revenue" if rev else "N/A"),
            ("EBITDA", ebitda, f"{ebitda/rev*100:.1f}% of Revenue" if rev else "N/A"),
            ("PBT", fd.profit_before_tax(), f"{fd.profit_before_tax()/rev*100:.1f}% of Revenue" if rev else "N/A"),
            ("PAT", fd.profit_after_tax(), f"{fd.profit_after_tax()/rev*100:.1f}% of Revenue" if rev else "N/A"),
        ]
        for lbl, amt, ratio in ratios:
            ws[f"A{r}"].value = "  " + lbl
            ws[f"C{r}"].value = amt
            ws[f"C{r}"].number_format = INR
            ws[f"C{r}"].alignment = _align("right")
            # ratio in column B area as comment
            ws[f"B{r}"].value = ratio
            ws[f"B{r}"].font = _font(size=8, color="595959")
            ws[f"B{r}"].alignment = _align("center", wrap=True)
            ws.column_dimensions["B"].width = 20
            r += 1

        ws.freeze_panes = "A6"

    # ── Projected P&L ─────────────────────────────────────────────────────────

    def _write_proj_pnl(self, pe: "ProjectionEngine") -> None:
        ws = self.wb.create_sheet("Projected P&L")
        fd = self.fd
        ws.column_dimensions["A"].width = 42
        for col, lbl in [("B", "Base Year"), ("C", "Year 1"), ("D", "Year 2"), ("E", "Year 3")]:
            ws.column_dimensions[col].width = 16

        r = 1
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value = f"{fd.company.upper()} — PROJECTED PROFIT & LOSS (3 YEARS)"
        ws[f"A{r}"].font = _font(bold=True, size=12, color=C_WHITE)
        ws[f"A{r}"].fill = _fill(C_HEADER_BG)
        ws[f"A{r}"].alignment = _align("center")
        r += 1

        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value = "Figures in ₹ | Projections are management estimates"
        ws[f"A{r}"].font = _font(size=9, color="595959")
        ws[f"A{r}"].alignment = _align("center")
        r += 1

        # Column headers
        years = [fd.period_label, "Year 1", "Year 2", "Year 3"]
        for i, col in enumerate(["A", "B", "C", "D", "E"]):
            ws[f"{col}{r}"].value = "Particulars" if i == 0 else years[i-1]
            ws[f"{col}{r}"].font = _font(bold=True, size=9, color=C_WHITE)
            ws[f"{col}{r}"].fill = _fill(C_MID_BLUE)
            ws[f"{col}{r}"].alignment = _align("center" if i > 0 else "left")
        r += 1

        pnl = pe.projected_pnl()  # list of 3 year dicts

        def proj_row(label: str, base_val: float, proj_vals: list[float],
                     bold=False, total=False, bg=None, fmt=INR):
            nonlocal r
            ws[f"A{r}"].value = label
            ws[f"A{r}"].font = _font(bold=bold or total)
            ws[f"B{r}"].value = base_val
            ws[f"B{r}"].number_format = fmt
            ws[f"B{r}"].alignment = _align("right")
            for i, (col, val) in enumerate(zip(["C", "D", "E"], proj_vals)):
                ws[f"{col}{r}"].value = val
                ws[f"{col}{r}"].number_format = fmt
                ws[f"{col}{r}"].font = _font(bold=bold or total)
                ws[f"{col}{r}"].alignment = _align("right")
            if total or bg:
                fill_c = bg or C_TOTAL_BG
                for col in ["A", "B", "C", "D", "E"]:
                    ws[f"{col}{r}"].fill = _fill(fill_c)
            r += 1

        base = fd  # actual year
        proj_row("Revenue from Operations",
                 base.revenue_from_ops(), [p["revenue"] for p in pnl], bold=True)
        proj_row("Other Income",
                 base.other_income(), [p["other_income"] for p in pnl])
        proj_row("TOTAL REVENUE",
                 base.revenue_from_ops() + base.other_income(),
                 [p["revenue"] + p["other_income"] for p in pnl],
                 total=True)
        r += 1
        proj_row("Cost of Materials / Purchases",
                 base.purchases(), [p["cogs"] for p in pnl])
        proj_row("Changes in Inventories",
                 base.opening_stock() - base.closing_stock(),
                 [p["stock_change"] for p in pnl])
        proj_row("Gross Profit",
                 base.revenue_from_ops() - base.purchases() + (base.closing_stock() - base.opening_stock()),
                 [p["gross_profit"] for p in pnl], bold=True, bg=C_LIGHT_BLUE)
        r += 1
        proj_row("Employee Benefits Expense",
                 base.employee_costs(), [p["employee"] for p in pnl])
        proj_row("Finance Costs",
                 base.finance_costs(), [p["finance"] for p in pnl])
        proj_row("Depreciation",
                 base.depreciation_from_fa(), [p["depreciation"] for p in pnl])
        proj_row("Other Expenses",
                 base.other_indirect_expenses() + base.direct_expenses(),
                 [p["other_expenses"] for p in pnl])
        proj_row("TOTAL EXPENSES",
                 base.total_expenses(),
                 [p["total_expenses"] for p in pnl], total=True)
        r += 1
        proj_row("EBITDA",
                 base.profit_before_tax() + base.finance_costs() + base.depreciation_from_fa(),
                 [p["ebitda"] for p in pnl], bold=True)
        proj_row("Profit Before Tax",
                 base.profit_before_tax(), [p["pbt"] for p in pnl], bold=True)
        proj_row("Tax Expense",
                 base.tax_expense(), [p["tax"] for p in pnl])
        proj_row("PROFIT AFTER TAX",
                 base.profit_after_tax(), [p["pat"] for p in pnl],
                 bold=True, bg=C_HEADER_BG)
        # Colour PAT row white text
        for col in ["A","B","C","D","E"]:
            ws[f"{col}{r-1}"].font = Font(bold=True, size=10, color=C_WHITE, name="Calibri")

        r += 2
        # Growth % rows
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value = "GROWTH & MARGIN METRICS"
        ws[f"A{r}"].font = _font(bold=True, size=9, color=C_WHITE)
        ws[f"A{r}"].fill = _fill(C_MID_BLUE)
        r += 1
        for metric, base_v, proj_vs in [
            ("Revenue Growth %", None, [p.get("revenue_growth") for p in pnl]),
            ("Gross Margin %", (base.revenue_from_ops() - base.purchases() + base.closing_stock() - base.opening_stock()) / base.revenue_from_ops() * 100 if base.revenue_from_ops() else 0,
             [p["gross_margin_pct"] for p in pnl]),
            ("EBITDA Margin %", (base.profit_before_tax() + base.finance_costs() + base.depreciation_from_fa()) / base.revenue_from_ops() * 100 if base.revenue_from_ops() else 0,
             [p["ebitda_margin"] for p in pnl]),
            ("PAT Margin %", base.profit_after_tax() / base.revenue_from_ops() * 100 if base.revenue_from_ops() else 0,
             [p["pat_margin"] for p in pnl]),
        ]:
            ws[f"A{r}"].value = "  " + metric
            ws[f"A{r}"].font = _font(size=9)
            if base_v is not None:
                ws[f"B{r}"].value = base_v; ws[f"B{r}"].number_format = "0.0%"; ws[f"B{r}"].value = base_v / 100
            for col, val in zip(["C","D","E"], proj_vs):
                if val is not None:
                    ws[f"{col}{r}"].value = val / 100
                    ws[f"{col}{r}"].number_format = "0.0%"
                    ws[f"{col}{r}"].alignment = _align("right")
            r += 1

        ws.freeze_panes = "A5"

    # ── Projected Balance Sheet ───────────────────────────────────────────────

    def _write_proj_bs(self, pe: "ProjectionEngine") -> None:
        ws = self.wb.create_sheet("Projected Balance Sheet")
        fd = self.fd
        ws.column_dimensions["A"].width = 46
        for col in ["B","C","D","E"]:
            ws.column_dimensions[col].width = 16

        r = 1
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value = f"{fd.company.upper()} — PROJECTED BALANCE SHEET (3 YEARS)"
        ws[f"A{r}"].font = _font(bold=True, size=12, color=C_WHITE)
        ws[f"A{r}"].fill = _fill(C_HEADER_BG)
        ws[f"A{r}"].alignment = _align("center")
        r += 1

        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value = "Provisional / Projected — For Discussion Purposes Only"
        ws[f"A{r}"].font = _font(size=9, color="FF0000")
        ws[f"A{r}"].alignment = _align("center")
        r += 1

        for i, (col, lbl) in enumerate(zip(["A","B","C","D","E"],
                                ["Particulars", "Base Year", "Year 1", "Year 2", "Year 3"])):
            ws[f"{col}{r}"].value = lbl
            ws[f"{col}{r}"].font = _font(bold=True, size=9, color=C_WHITE)
            ws[f"{col}{r}"].fill = _fill(C_MID_BLUE)
            ws[f"{col}{r}"].alignment = _align("left" if i == 0 else "center")
        r += 1

        bs = pe.projected_bs()

        def bs_row(label, base_val, proj_vals, bold=False, total=False, bg=None):
            nonlocal r
            ws[f"A{r}"].value = label
            ws[f"A{r}"].font = _font(bold=bold or total)
            ws[f"B{r}"].value = base_val; ws[f"B{r}"].number_format = INR
            ws[f"B{r}"].alignment = _align("right")
            for col, val in zip(["C","D","E"], proj_vals):
                ws[f"{col}{r}"].value = val
                ws[f"{col}{r}"].number_format = INR
                ws[f"{col}{r}"].font = _font(bold=bold or total)
                ws[f"{col}{r}"].alignment = _align("right")
            if total or bg:
                for col in ["A","B","C","D","E"]:
                    ws[f"{col}{r}"].fill = _fill(bg or C_TOTAL_BG)
            r += 1

        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value = "EQUITY & LIABILITIES"; ws[f"A{r}"].font = _font(bold=True); ws[f"A{r}"].fill = _fill(C_SUBHD_BG); r += 1
        bs_row("  Share Capital",         fd.share_capital(),         [b["share_capital"] for b in bs])
        bs_row("  Reserves & Surplus",    fd.reserves_surplus(),      [b["reserves"] for b in bs], bold=True)
        bs_row("Total Equity",
               fd.total_equity(),         [b["total_equity"] for b in bs], total=True)
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value = "Non-Current Liabilities"; ws[f"A{r}"].font = _font(bold=True); ws[f"A{r}"].fill = _fill(C_SUBHD_BG); r += 1
        bs_row("  Long-Term Borrowings",  fd.long_term_borrowings(),  [b["lt_borrowings"] for b in bs])
        bs_row("  Deferred Tax Liability",fd.deferred_tax_liability(),[b["dtl"] for b in bs])
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value = "Current Liabilities"; ws[f"A{r}"].font = _font(bold=True); ws[f"A{r}"].fill = _fill(C_SUBHD_BG); r += 1
        bs_row("  Short-Term Borrowings", fd.short_term_borrowings(), [b["st_borrowings"] for b in bs])
        bs_row("  Trade Payables",        fd.trade_payables(),        [b["trade_payables"] for b in bs])
        bs_row("  Other Current Liab.",   fd.other_current_liabilities() + max(0, fd.duties_and_taxes_net()),
                                          [b["other_cl"] for b in bs])
        bs_row("  Short-Term Provisions", fd.short_term_provisions(), [b["provisions"] for b in bs])
        bs_row("TOTAL EQUITY & LIABILITIES",
               fd.total_equity_liabilities(), [b["total_el"] for b in bs],
               bold=True, bg=C_HEADER_BG)
        for col in ["A","B","C","D","E"]:
            ws[f"{col}{r-1}"].font = Font(bold=True, size=10, color=C_WHITE, name="Calibri")

        r += 2
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value = "ASSETS"; ws[f"A{r}"].font = _font(bold=True); ws[f"A{r}"].fill = _fill(C_SUBHD_BG); r += 1
        bs_row("  Fixed Assets (Net)",    fd.net_fixed_assets(),      [b["fixed_assets"] for b in bs])
        bs_row("  Long-Term Loans & Adv.",fd.long_term_loans_advances(),[b["lt_loans"] for b in bs])
        bs_row("Total Non-Current Assets",
               fd.total_noncurrent_assets(), [b["total_nca"] for b in bs], total=True)
        r += 1
        bs_row("  Inventories",           fd.closing_stock(),         [b["inventory"] for b in bs])
        bs_row("  Trade Receivables",     fd.trade_receivables(),     [b["debtors"] for b in bs])
        bs_row("  Cash & Cash Equiv.",    fd.cash_and_bank(),         [b["cash"] for b in bs])
        bs_row("  Other Current Assets",  fd.other_current_assets(),  [b["other_ca"] for b in bs])
        bs_row("Total Current Assets",
               fd.total_current_assets(), [b["total_ca"] for b in bs], total=True)
        r += 1
        bs_row("TOTAL ASSETS",
               fd.total_assets(),         [b["total_assets"] for b in bs],
               bold=True, bg=C_HEADER_BG)
        for col in ["A","B","C","D","E"]:
            ws[f"{col}{r-1}"].font = Font(bold=True, size=10, color=C_WHITE, name="Calibri")

        ws.freeze_panes = "A5"

    # ── Assumptions sheet ─────────────────────────────────────────────────────

    def _write_assumptions(self) -> None:
        ws = self.wb.create_sheet("Assumptions")
        p = self.proj
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 18

        ws.merge_cells("A1:B1")
        ws["A1"].value = "PROJECTION ASSUMPTIONS"
        ws["A1"].font = _font(bold=True, size=12, color=C_WHITE)
        ws["A1"].fill = _fill(C_HEADER_BG)
        ws["A1"].alignment = _align("center")

        rows = [
            ("Revenue Growth – Year 1",         f"{p.rev_growth_y1:.1f}%"),
            ("Revenue Growth – Year 2",         f"{p.rev_growth_y2:.1f}%"),
            ("Revenue Growth – Year 3",         f"{p.rev_growth_y3:.1f}%"),
            ("Gross Margin %",                  f"{p.gross_margin_pct:.1f}%"),
            ("Operating Expenses Growth %",     f"{p.opex_growth_pct:.1f}%"),
            ("Inventory / COGS Days",           f"{p.inventory_days:.0f} days"),
            ("Debtor Days",                     f"{p.debtor_days:.0f} days"),
            ("Creditor Days",                   f"{p.creditor_days:.0f} days"),
            ("Annual Loan Repayment (₹)",       f"₹ {p.loan_repayment_pa:,.0f}"),
            ("New Borrowings per Year (₹)",     f"₹ {p.new_borrowings_pa:,.0f}"),
            ("Interest Rate %",                 f"{p.interest_rate_pct:.1f}%"),
            ("Capital Expenditure / Year (₹)",  f"₹ {p.capex_pa:,.0f}"),
            ("Depreciation Rate (WDV) %",       f"{p.depreciation_rate_pct:.1f}%"),
            ("Effective Tax Rate %",            f"{p.tax_rate_pct:.1f}%"),
            ("Other Income (fixed ₹ / year)",   f"₹ {p.other_income_pa:,.0f}"),
        ]
        for i, (lbl, val) in enumerate(rows, start=2):
            ws[f"A{i}"].value = lbl
            ws[f"A{i}"].font = _font(size=10)
            ws[f"B{i}"].value = val
            ws[f"B{i}"].font = _font(size=10, bold=True)
            ws[f"B{i}"].alignment = _align("right")
            if i % 2 == 0:
                ws[f"A{i}"].fill = _fill(C_LIGHT_BLUE)
                ws[f"B{i}"].fill = _fill(C_LIGHT_BLUE)

        ws[f"A{len(rows)+3}"].value = "Generated on"
        ws[f"B{len(rows)+3}"].value = date.today().strftime("%d %B %Y")
        ws[f"A{len(rows)+3}"].font = _font(size=9, color="595959")
        ws[f"B{len(rows)+3}"].font = _font(size=9, color="595959")


# ─── Projection Engine ────────────────────────────────────────────────────────

@dataclass
class ProjectionInputs:
    rev_growth_y1:       float = 15.0   # %
    rev_growth_y2:       float = 15.0   # %
    rev_growth_y3:       float = 10.0   # %
    gross_margin_pct:    float = 30.0   # % of revenue
    opex_growth_pct:     float = 10.0   # % growth in operating expenses
    inventory_days:      float = 45.0   # days
    debtor_days:         float = 60.0   # days
    creditor_days:       float = 45.0   # days
    loan_repayment_pa:   float = 0.0    # ₹ per year
    new_borrowings_pa:   float = 0.0    # ₹ per year
    interest_rate_pct:   float = 14.0   # % on average borrowings
    capex_pa:            float = 0.0    # ₹ per year
    depreciation_rate_pct: float = 15.0 # % WDV
    tax_rate_pct:        float = 25.0   # %
    other_income_pa:     float = 0.0    # ₹ fixed per year


class ProjectionEngine:
    def __init__(self, fd: FinancialData, p: ProjectionInputs):
        self.fd = fd
        self.p  = p

    def projected_pnl(self) -> list[dict]:
        fd = self.fd
        p  = self.p
        base_rev  = fd.revenue_from_ops()
        base_opex = (fd.employee_costs()
                     + fd.other_indirect_expenses()
                     + fd.direct_expenses())

        results = []
        prev_rev = base_rev
        prev_rev_growth = None
        growth_rates = [p.rev_growth_y1, p.rev_growth_y2, p.rev_growth_y3]

        for yr_idx, g in enumerate(growth_rates):
            rev = prev_rev * (1 + g / 100)
            cogs = rev * (1 - p.gross_margin_pct / 100)
            stock_change = -(rev / 365 * p.inventory_days - fd.closing_stock()) if yr_idx == 0 else 0
            gross_profit = rev - cogs + stock_change

            opex_mult = (1 + p.opex_growth_pct / 100) ** (yr_idx + 1)
            employee  = fd.employee_costs() * opex_mult
            other_exp = (fd.other_indirect_expenses() + fd.direct_expenses()) * opex_mult

            # Borrowings for interest calculation
            lt_borr = max(0, fd.long_term_borrowings()
                          + (p.new_borrowings_pa - p.loan_repayment_pa) * (yr_idx + 1))
            st_borr = fd.short_term_borrowings()
            avg_borr = lt_borr + st_borr
            finance  = avg_borr * p.interest_rate_pct / 100

            # Depreciation (WDV)
            if yr_idx == 0:
                fa_base = fd.net_fixed_assets() + p.capex_pa
            else:
                fa_base = results[-1]["net_fa"]
            depr = fa_base * p.depreciation_rate_pct / 100
            net_fa = fa_base - depr + p.capex_pa

            total_exp = cogs + employee + finance + depr + other_exp
            ebitda    = rev - cogs - employee - other_exp
            pbt       = rev - total_exp + p.other_income_pa
            tax       = max(0, pbt * p.tax_rate_pct / 100)
            pat       = pbt - tax

            results.append({
                "revenue":          rev,
                "revenue_growth":   g,
                "other_income":     p.other_income_pa,
                "cogs":             cogs,
                "stock_change":     stock_change,
                "gross_profit":     gross_profit,
                "gross_margin_pct": p.gross_margin_pct,
                "employee":         employee,
                "finance":          finance,
                "depreciation":     depr,
                "other_expenses":   other_exp,
                "total_expenses":   total_exp,
                "ebitda":           ebitda,
                "ebitda_margin":    ebitda / rev * 100 if rev else 0,
                "pbt":              pbt,
                "tax":              tax,
                "pat":              pat,
                "pat_margin":       pat / rev * 100 if rev else 0,
                "net_fa":           net_fa,
            })
            prev_rev = rev

        return results

    def projected_bs(self) -> list[dict]:
        fd = self.fd
        p  = self.p
        pnl = self.projected_pnl()

        bs_list = []
        prev_reserves = fd.reserves_surplus()
        prev_lt_borr  = fd.long_term_borrowings()
        prev_fa       = fd.net_fixed_assets()

        for yr_idx, yr in enumerate(pnl):
            rev  = yr["revenue"]
            cogs = yr["cogs"]
            pat  = yr["pat"]

            # Equity
            reserves    = prev_reserves + pat
            share_cap   = fd.share_capital()
            total_equity = share_cap + reserves

            # Borrowings
            lt_borr = max(0, prev_lt_borr + p.new_borrowings_pa - p.loan_repayment_pa)
            st_borr = fd.short_term_borrowings()

            # Working capital
            inventory = cogs / 365 * p.inventory_days
            debtors   = rev  / 365 * p.debtor_days
            creditors = cogs / 365 * p.creditor_days

            # Fixed assets
            fa = yr["net_fa"]

            # Other items (kept at base year)
            lt_loans = fd.long_term_loans_advances()
            other_ca = fd.other_current_assets()
            provisions = fd.short_term_provisions()
            dtl       = fd.deferred_tax_liability()

            # Total liabilities
            other_cl = fd.other_current_liabilities()
            total_el = total_equity + lt_borr + dtl + st_borr + creditors + other_cl + provisions

            # Cash (plug — balances the sheet)
            total_nca  = fa + lt_loans
            total_excl_cash = total_nca + inventory + debtors + other_ca
            cash = total_el - total_excl_cash

            total_ca = inventory + debtors + cash + other_ca
            total_assets = total_nca + total_ca

            bs_list.append({
                "share_capital":  share_cap,
                "reserves":       reserves,
                "total_equity":   total_equity,
                "lt_borrowings":  lt_borr,
                "dtl":            dtl,
                "st_borrowings":  st_borr,
                "trade_payables": creditors,
                "other_cl":       other_cl,
                "provisions":     provisions,
                "total_el":       total_el,
                "fixed_assets":   fa,
                "lt_loans":       lt_loans,
                "total_nca":      total_nca,
                "inventory":      inventory,
                "debtors":        debtors,
                "cash":           cash,
                "other_ca":       other_ca,
                "total_ca":       total_ca,
                "total_assets":   total_assets,
            })
            prev_reserves = reserves
            prev_lt_borr  = lt_borr
            prev_fa       = fa

        return bs_list


# ─── Tkinter UI ───────────────────────────────────────────────────────────────

class App:
    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title("Tally Financial Statements Generator")
        self.root.geometry("780x680")
        self.root.minsize(700, 600)
        self.root.configure(bg="#F0F4F8")

        self.db_path     = tk.StringVar()
        self.out_dir     = tk.StringVar(value=str(Path.home() / "Desktop"))
        self.op_stock    = tk.StringVar()
        self.cl_stock    = tk.StringVar()
        self.status_var  = tk.StringVar(value="Select a Tally SQLite file to begin.")
        self.fd: FinancialData | None = None
        self.reclassify_map: dict[str, str] = {}
        self._mapping_vars: dict[str, tk.StringVar] = {}  # primary_group → StringVar for dropdown

        # Projection inputs
        self.proj_vars: dict[str, tk.StringVar] = {
            "rev_growth_y1":       tk.StringVar(value="15"),
            "rev_growth_y2":       tk.StringVar(value="15"),
            "rev_growth_y3":       tk.StringVar(value="10"),
            "gross_margin_pct":    tk.StringVar(value="30"),
            "opex_growth_pct":     tk.StringVar(value="10"),
            "inventory_days":      tk.StringVar(value="45"),
            "debtor_days":         tk.StringVar(value="60"),
            "creditor_days":       tk.StringVar(value="45"),
            "loan_repayment_pa":   tk.StringVar(value="0"),
            "new_borrowings_pa":   tk.StringVar(value="0"),
            "interest_rate_pct":   tk.StringVar(value="14"),
            "capex_pa":            tk.StringVar(value="0"),
            "depreciation_rate_pct": tk.StringVar(value="15"),
            "tax_rate_pct":        tk.StringVar(value="25"),
            "other_income_pa":     tk.StringVar(value="0"),
        }

        self._build()
        self.root.mainloop()

    def _build(self) -> None:
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TNotebook",        background="#F0F4F8")
        style.configure("TNotebook.Tab",    padding=[12, 6], font=("Calibri", 10, "bold"))
        style.configure("TFrame",           background="#F0F4F8")
        style.configure("TLabel",           background="#F0F4F8", font=("Calibri", 10))
        style.configure("TButton",          font=("Calibri", 10, "bold"), padding=6)
        style.configure("Accent.TButton",   font=("Calibri", 11, "bold"), padding=8)
        style.configure("TEntry",           font=("Calibri", 10))
        style.configure("TLabelframe",      background="#F0F4F8",
                         font=("Calibri", 10, "bold"))
        style.configure("TLabelframe.Label",background="#F0F4F8",
                         font=("Calibri", 10, "bold"))

        # ── Title bar ────────────────────────────────────────────────────────
        title_frame = tk.Frame(self.root, bg="#1F3864", pady=14)
        title_frame.pack(fill="x")
        tk.Label(title_frame,
                 text="Tally Financial Statements Generator",
                 bg="#1F3864", fg="white",
                 font=("Calibri", 15, "bold")).pack()
        tk.Label(title_frame,
                 text="Schedule III Balance Sheet  •  P&L  •  3-Year Projections",
                 bg="#1F3864", fg="#BDD7EE",
                 font=("Calibri", 10)).pack()

        main = ttk.Frame(self.root, padding=16)
        main.pack(fill="both", expand=True)

        # ── File selection ────────────────────────────────────────────────────
        file_frame = ttk.LabelFrame(main, text="  Database File", padding=10)
        file_frame.pack(fill="x", pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)

        ttk.Label(file_frame, text="Tally SQLite:").grid(row=0, column=0, sticky="w", padx=(0,8))
        ttk.Entry(file_frame, textvariable=self.db_path, width=55).grid(row=0, column=1, sticky="ew")
        ttk.Button(file_frame, text="Browse…", command=self._pick_db).grid(row=0, column=2, padx=(8,0))

        ttk.Label(file_frame, text="Output Folder:").grid(row=1, column=0, sticky="w", padx=(0,8), pady=(6,0))
        ttk.Entry(file_frame, textvariable=self.out_dir, width=55).grid(row=1, column=1, sticky="ew", pady=(6,0))
        ttk.Button(file_frame, text="Browse…", command=self._pick_outdir).grid(row=1, column=2, padx=(8,0), pady=(6,0))

        # ── Tabs ─────────────────────────────────────────────────────────────
        self.nb = ttk.Notebook(main)
        self.nb.pack(fill="both", expand=True, pady=(0,10))

        self._build_actual_tab()
        self._build_proj_tab()
        self._build_validation_tab()
        self._build_mapping_tab()

        # ── Status + action ───────────────────────────────────────────────────
        bottom = ttk.Frame(main)
        bottom.pack(fill="x")
        self.status_lbl = ttk.Label(bottom, textvariable=self.status_var,
                                    foreground="#2F5496")
        self.status_lbl.pack(side="left", fill="x", expand=True)
        ttk.Button(bottom, text="❌ Clear", command=self._clear,
                   style="TButton").pack(side="right", padx=(8,0))

    def _build_actual_tab(self) -> None:
        tab = ttk.Frame(self.nb, padding=12)
        self.nb.add(tab, text="Actual Statements")
        tab.columnconfigure(1, weight=1)

        # Stock adjustments
        stock_frame = ttk.LabelFrame(tab, text="  Stock Values", padding=10)
        stock_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0,12))
        stock_frame.columnconfigure(1, weight=1)

        help_text = ("Tally's 'Closing Stock' ledger holds accounting-entry stock values.\n"
                     "Enter corrected values below if they differ from the audit trial balance.\n"
                     "Leave blank to use Tally's own figures.")
        ttk.Label(stock_frame, text=help_text, foreground="#595959",
                  font=("Calibri", 9)).grid(row=0, column=0, columnspan=3,
                                             sticky="w", pady=(0,8))

        ttk.Label(stock_frame, text="Opening Stock (₹):").grid(row=1, column=0, sticky="w", padx=(0,10))
        self.op_entry = ttk.Entry(stock_frame, textvariable=self.op_stock, width=20)
        self.op_entry.grid(row=1, column=1, sticky="w")
        ttk.Label(stock_frame, text="(from Tally if blank)",
                  foreground="#888", font=("Calibri",9)).grid(row=1, column=2, sticky="w", padx=8)

        ttk.Label(stock_frame, text="Closing Stock (₹):").grid(row=2, column=0, sticky="w", padx=(0,10), pady=(6,0))
        self.cl_entry = ttk.Entry(stock_frame, textvariable=self.cl_stock, width=20)
        self.cl_entry.grid(row=2, column=1, sticky="w", pady=(6,0))
        ttk.Label(stock_frame, text="(from Tally if blank)",
                  foreground="#888", font=("Calibri",9)).grid(row=2, column=2, sticky="w", padx=8, pady=(6,0))

        ttk.Button(stock_frame, text="Apply & Preview Numbers",
                   command=self._preview_actual).grid(row=3, column=0, columnspan=3,
                                                       pady=(10,0), sticky="w")

        # Preview panel
        self.preview_text = tk.Text(tab, height=16, width=72, font=("Courier", 9),
                                    bg="#FAFAFA", relief="flat", borderwidth=1)
        self.preview_text.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0,8))
        tab.rowconfigure(1, weight=1)

        btn_frame = ttk.Frame(tab)
        btn_frame.grid(row=2, column=0, columnspan=2, sticky="ew")
        ttk.Button(btn_frame, text="Generate Actual Statements (Excel)",
                   command=self._gen_actual, style="Accent.TButton").pack(side="right")

    def _build_proj_tab(self) -> None:
        tab = ttk.Frame(self.nb, padding=12)
        self.nb.add(tab, text="3-Year Projections")

        canvas = tk.Canvas(tab, bg="#F0F4F8", highlightthickness=0)
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = ttk.Frame(canvas, padding=8)
        canvas_window = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(canvas_window, width=event.width if event.width > 1 else canvas.winfo_width())
        inner.bind("<Configure>", _on_configure)
        canvas.bind("<Configure>", _on_configure)

        inner.columnconfigure(1, weight=1)
        inner.columnconfigure(3, weight=1)

        fields = [
            ("Revenue Growth – Year 1 (%)",            "rev_growth_y1"),
            ("Revenue Growth – Year 2 (%)",            "rev_growth_y2"),
            ("Revenue Growth – Year 3 (%)",            "rev_growth_y3"),
            ("Gross Margin % (Revenue − COGS)",        "gross_margin_pct"),
            ("Operating Expense Growth (% per year)",  "opex_growth_pct"),
            ("Inventory Days (Stock / COGS × 365)",    "inventory_days"),
            ("Debtor Days (Receivables / Rev × 365)",  "debtor_days"),
            ("Creditor Days (Payables / COGS × 365)",  "creditor_days"),
            ("Loan Repayment per Year (₹)",            "loan_repayment_pa"),
            ("New Borrowings per Year (₹)",            "new_borrowings_pa"),
            ("Interest Rate on Borrowings (%)",        "interest_rate_pct"),
            ("Capital Expenditure per Year (₹)",       "capex_pa"),
            ("Depreciation Rate – WDV (%)",            "depreciation_rate_pct"),
            ("Effective Tax Rate (%)",                  "tax_rate_pct"),
            ("Other / Fixed Income per Year (₹)",      "other_income_pa"),
        ]

        hints = {
            "gross_margin_pct":  "Gross Profit as % of revenue (Revenue − Purchases ± Stock)",
            "debtor_days":       "Days outstanding for trade receivables",
            "creditor_days":     "Days outstanding for trade payables",
            "inventory_days":    "Days of stock held (based on COGS)",
            "interest_rate_pct": "Average annualised interest on total borrowings",
            "depreciation_rate_pct": "Written-down-value rate; 15% is typical for plant & equipment",
            "tax_rate_pct":      "Effective corporate tax rate (base rate 22% + surcharge ≈ 25.17%)",
        }

        for i, (label, key) in enumerate(fields):
            row_i = i // 2
            col_base = (i % 2) * 2
            ttk.Label(inner, text=label, font=("Calibri", 9)).grid(
                row=row_i, column=col_base, sticky="w", padx=(0,8), pady=4)
            e = ttk.Entry(inner, textvariable=self.proj_vars[key], width=14)
            e.grid(row=row_i, column=col_base + 1, sticky="ew", pady=4)
            if key in hints:
                e.bind("<FocusIn>", lambda ev, h=hints[key]: self._set_status(h))

        hint_r = len(fields) // 2 + 1
        hint_box = tk.Text(inner, height=5, font=("Calibri", 9), bg="#FFF9E6",
                           relief="flat", wrap="word", borderwidth=1)
        hint_box.insert("1.0",
            "GUIDANCE:\n"
            "• Gross Margin: for a trading company like Fortius, typical is 20–35%.\n"
            "• Debtor Days: if your debtors turn over in 60 days, enter 60.\n"
            "• Creditor Days: industry avg for electronics trading ≈ 30–45 days.\n"
            "• New Borrowings / Repayments: enter absolute ₹ values per year.\n"
            "• Leave CapEx at 0 if no new investments planned.\n"
            "• The projected Balance Sheet uses a cash plug to balance automatically.")
        hint_box.config(state="disabled")
        hint_box.grid(row=hint_r, column=0, columnspan=4, sticky="ew",
                      pady=(12,8), padx=4)

        btn_frame = ttk.Frame(inner)
        btn_frame.grid(row=hint_r + 1, column=0, columnspan=4, sticky="e", pady=8)
        ttk.Button(btn_frame, text="Generate Projected Statements + Actual (All Sheets)",
                   command=self._gen_all, style="Accent.TButton").pack(side="right")
        ttk.Button(btn_frame, text="Projections Only",
                   command=self._gen_proj_only, style="TButton").pack(side="right", padx=(0,8))

    def _build_validation_tab(self) -> None:
        tab = ttk.Frame(self.nb, padding=8)
        self.nb.add(tab, text="Validation")
        tab.rowconfigure(1, weight=1)
        tab.columnconfigure(0, weight=1)

        top = ttk.Frame(tab)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        self.val_summary = ttk.Label(top,
            text="Load a database file to see validation results.",
            font=("Calibri", 10, "bold"), foreground="#2F5496")
        self.val_summary.pack(side="left")
        ttk.Button(top, text="Re-run Validation",
                   command=self._run_validation).pack(side="right")

        cols = ("Severity", "Category", "Message", "Detail")
        tree = ttk.Treeview(tab, columns=cols, show="headings", height=20)
        tree.heading("Severity",  text="Severity",  anchor="center")
        tree.heading("Category",  text="Category",  anchor="w")
        tree.heading("Message",   text="Message",   anchor="w")
        tree.heading("Detail",    text="Detail",    anchor="w")
        tree.column("Severity",  width=80,  stretch=False, anchor="center")
        tree.column("Category",  width=130, stretch=False)
        tree.column("Message",   width=320)
        tree.column("Detail",    width=380)

        tree.tag_configure("ERROR",   background="#FFCCCC", foreground="#CC0000")
        tree.tag_configure("WARNING", background="#FFF2CC", foreground="#7B5800")
        tree.tag_configure("INFO",    background="#E8F5E9", foreground="#1B5E20")

        vsb = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=1, column=0, sticky="nsew")
        vsb.grid(row=1, column=1, sticky="ns")

        self.val_tree = tree

    def _build_mapping_tab(self) -> None:
        tab = ttk.Frame(self.nb, padding=8)
        self.nb.add(tab, text="Group Mapping")
        tab.rowconfigure(1, weight=1)
        tab.columnconfigure(0, weight=1)

        info = ttk.Label(tab,
            text="Assign unrecognised primary groups to standard Schedule III heads.\n"
                 "Groups already in the built-in map are shown but cannot be changed here.",
            font=("Calibri", 9), foreground="#595959")
        info.grid(row=0, column=0, sticky="ew", pady=(0,6))

        # Scrollable container for rows
        container = ttk.Frame(tab)
        container.grid(row=1, column=0, sticky="nsew")
        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)

        canvas = tk.Canvas(container, bg="#F0F4F8", highlightthickness=0)
        vsb = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        self._mapping_inner = ttk.Frame(canvas, padding=4)
        self._mapping_canvas_id = canvas.create_window((0, 0), window=self._mapping_inner, anchor="nw")

        def _on_inner_resize(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(self._mapping_canvas_id, width=canvas.winfo_width())
        self._mapping_inner.bind("<Configure>", _on_inner_resize)
        canvas.bind("<Configure>", _on_inner_resize)
        self._mapping_canvas = canvas

        btn_row = ttk.Frame(tab)
        btn_row.grid(row=2, column=0, sticky="ew", pady=(6,0))
        ttk.Button(btn_row, text="Apply Mapping & Reload Preview",
                   command=self._apply_mapping).pack(side="right")
        self._mapping_vars = {}
        self._mapping_inner_rows = []   # list of (frame) for clearing

    def _populate_mapping_tab(self, fd: "FinancialData") -> None:
        """Fill the mapping tab from the loaded FinancialData."""
        from collections import defaultdict
        pg_balance: dict[str, float] = defaultdict(float)
        pg_parents: dict[str, set[str]] = defaultdict(set)
        for l in fd.ledgers:
            pg_balance[l.primary_group] += l.closing
            pg_parents[l.primary_group].add(l.parent)

        # Clear old rows
        for child in self._mapping_inner.winfo_children():
            child.destroy()
        self._mapping_vars.clear()

        ALL_STANDARD = sorted(set(BS_MAP.keys()) | set(PNL_MAP.keys()))
        DROPDOWN_VALUES = ["(Use Auto-Infer)"] + ALL_STANDARD

        # Column headers
        for c, lbl in enumerate(["Primary Group", "Net Balance (Rs.)", "Auto-Inferred Head", "Your Override"]):
            ttk.Label(self._mapping_inner, text=lbl, font=("Calibri", 9, "bold"),
                      background="#BDD7EE").grid(row=0, column=c, sticky="ew", padx=2, pady=2)
        self._mapping_inner.columnconfigure(0, weight=1)
        self._mapping_inner.columnconfigure(2, weight=1)
        self._mapping_inner.columnconfigure(3, weight=1)

        sorted_pgs = sorted(pg_balance.items(), key=lambda x: abs(x[1]), reverse=True)
        for row_i, (pg, bal) in enumerate(sorted_pgs, start=1):
            in_map = pg in BS_MAP or pg in PNL_MAP
            auto = pg if in_map else (infer_standard_group(pg, pg_parents[pg]) or "(no match)")
            current_override = self.reclassify_map.get(pg, "(Use Auto-Infer)")

            bg = "#F0F4F8" if row_i % 2 == 0 else "#FFFFFF"
            ttk.Label(self._mapping_inner, text=pg,
                      font=("Calibri", 9, "bold" if not in_map else "normal"),
                      background=bg, foreground="#CC0000" if not in_map else "#333333"
                      ).grid(row=row_i, column=0, sticky="ew", padx=2, pady=1)
            ttk.Label(self._mapping_inner, text=f"{bal:,.0f}",
                      font=("Calibri", 9), background=bg
                      ).grid(row=row_i, column=1, sticky="e", padx=2, pady=1)
            ttk.Label(self._mapping_inner, text=auto,
                      font=("Calibri", 9, "italic"), foreground="#595959", background=bg
                      ).grid(row=row_i, column=2, sticky="ew", padx=2, pady=1)

            if in_map:
                ttk.Label(self._mapping_inner, text="(built-in)",
                          font=("Calibri", 9), foreground="#888", background=bg
                          ).grid(row=row_i, column=3, sticky="ew", padx=2, pady=1)
            else:
                var = tk.StringVar(value=current_override)
                self._mapping_vars[pg] = var
                cb = ttk.Combobox(self._mapping_inner, textvariable=var,
                                  values=DROPDOWN_VALUES, state="readonly", width=28)
                cb.grid(row=row_i, column=3, sticky="ew", padx=2, pady=1)

    def _apply_mapping(self) -> None:
        """Save combobox selections to reclassify_map and reload preview."""
        new_map = {}
        for pg, var in self._mapping_vars.items():
            val = var.get()
            if val and val != "(Use Auto-Infer)":
                new_map[pg] = val
        self.reclassify_map = new_map
        if self.db_path.get():
            self._preview_actual()
            self._run_validation()

    # ── Actions ───────────────────────────────────────────────────────────────

    def _set_status(self, msg: str) -> None:
        self.status_var.set(msg)

    def _pick_db(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Tally SQLite File",
            filetypes=[("SQLite Database", "*.sqlite *.db"), ("All Files", "*.*")]
        )
        if path:
            self.db_path.set(path)
            self._load_db()

    def _pick_outdir(self) -> None:
        d = filedialog.askdirectory(title="Select Output Folder")
        if d:
            self.out_dir.set(d)

    def _load_db(self) -> None:
        path = self.db_path.get()
        if not path or not os.path.isfile(path):
            return
        try:
            self.fd = load_from_sqlite(
                path,
                opening_stock_override=self._parse_stock(self.op_stock.get()),
                closing_stock_override=self._parse_stock(self.cl_stock.get()),
            )
            self._set_status(
                f"Loaded: {self.fd.company}  |  {self.fd.period_from} → {self.fd.period_to}"
            )
            self._preview_actual()
            self._run_validation()
            self._populate_mapping_tab(self.fd)
        except Exception as e:
            messagebox.showerror("Load Error", str(e))

    def _run_validation(self) -> None:
        fd = self.fd
        if fd is None:
            return
        vr = validate_financial_data(fd)
        # Populate treeview
        for item in self.val_tree.get_children():
            self.val_tree.delete(item)
        for chk in vr.checks:
            self.val_tree.insert("", "end",
                values=(chk.severity, chk.category, chk.message, chk.detail),
                tags=(chk.severity,))
        # Update summary label
        summary = vr.summary()
        color = "#CC0000" if vr.has_errors else ("#7B5800" if vr.warnings else "#1B5E20")
        self.val_summary.configure(text=summary, foreground=color)

    def _parse_stock(self, val: str) -> float | None:
        val = val.strip().replace(",", "")
        try:
            return float(val) if val else None
        except ValueError:
            return None

    def _get_fd(self) -> FinancialData | None:
        """Return (possibly refreshed) FinancialData."""
        path = self.db_path.get()
        if not path or not os.path.isfile(path):
            messagebox.showwarning("No File", "Please select a Tally SQLite file first.")
            return None
        try:
            fd = load_from_sqlite(
                path,
                opening_stock_override=self._parse_stock(self.op_stock.get()),
                closing_stock_override=self._parse_stock(self.cl_stock.get()),
                reclassify_map=self.reclassify_map,
            )
            self.fd = fd
            return fd
        except Exception as e:
            messagebox.showerror("Load Error", str(e))
            return None

    def _preview_actual(self) -> None:
        fd = self._get_fd()
        if fd is None:
            return
        self.preview_text.config(state="normal")
        self.preview_text.delete("1.0", "end")

        def line(label, amount, indent=0):
            prefix = "  " * indent
            self.preview_text.insert("end", f"{prefix}{label:<45} {amount:>16,.2f}\n")

        def divider():
            self.preview_text.insert("end", "─" * 64 + "\n")

        self.preview_text.insert("end",
            f"  {fd.company}\n"
            f"  Balance Sheet Preview — as at {fd.period_label}\n"
        )
        divider()
        self.preview_text.insert("end", "  EQUITY & LIABILITIES\n")
        line("Share Capital", fd.share_capital(), 1)
        line("Reserves & Surplus", fd.reserves_surplus(), 1)
        line("Long-Term Borrowings", fd.long_term_borrowings(), 1)
        line("Deferred Tax Liability", fd.deferred_tax_liability(), 1)
        line("Short-Term Borrowings (OD/CC)", fd.short_term_borrowings(), 1)
        line("Trade Payables", fd.trade_payables(), 1)
        line("Duties & Taxes (Net payable)", max(0, fd.duties_and_taxes_net()), 1)
        line("Other Current Liabilities", fd.other_current_liabilities(), 1)
        line("Short-Term Provisions", fd.short_term_provisions(), 1)
        divider()
        line("TOTAL EQUITY & LIABILITIES", fd.total_equity_liabilities(), 0)
        divider()
        self.preview_text.insert("end", "\n  ASSETS\n")
        line("Fixed Assets (Net)", fd.net_fixed_assets(), 1)
        line("Long-Term Loans & Advances", fd.long_term_loans_advances(), 1)
        line("Closing Stock (Inventories)", fd.closing_stock(), 1)
        line("Trade Receivables", fd.trade_receivables(), 1)
        line("Cash & Cash Equivalents", fd.cash_and_bank(), 1)
        line("Duties & Taxes Refund (GST/TDS)", fd.duties_and_taxes_asset(), 1)
        line("Other Current Assets", fd.other_current_assets(), 1)
        divider()
        line("TOTAL ASSETS", fd.total_assets(), 0)
        diff = fd.total_assets() - fd.total_equity_liabilities()
        divider()
        self.preview_text.insert("end", f"\n  Difference (Assets − E&L): {diff:,.2f}")
        if abs(diff) > 10:
            self.preview_text.insert("end", "  ← REVIEW REQUIRED\n")
        else:
            self.preview_text.insert("end", "  ← OK\n")

        self.preview_text.insert("end",
            f"\n  P&L SUMMARY\n"
            f"  Revenue from Operations: {fd.revenue_from_ops():>20,.2f}\n"
            f"  Cost of Materials:       {fd.purchases():>20,.2f}\n"
            f"  Gross Profit:            {fd.revenue_from_ops()-fd.purchases()+(fd.closing_stock()-fd.opening_stock()):>20,.2f}\n"
            f"  Finance Costs:           {fd.finance_costs():>20,.2f}\n"
            f"  Profit Before Tax:       {fd.profit_before_tax():>20,.2f}\n"
            f"  Profit After Tax:        {fd.profit_after_tax():>20,.2f}\n"
        )
        self.preview_text.config(state="disabled")

    def _get_proj_inputs(self) -> ProjectionInputs | None:
        try:
            return ProjectionInputs(**{
                k: float(v.get().replace(",",""))
                for k, v in self.proj_vars.items()
            })
        except ValueError as e:
            messagebox.showerror("Invalid Input",
                                 f"Please enter valid numbers in the projection fields.\n{e}")
            return None

    def _output_path(self, suffix="") -> str:
        fd = self.fd
        company_short = (fd.company.split()[0] if fd else "Company").replace(",","").replace(".","")
        year = fd.period_to[:4] if fd else "2026"
        fname = f"{company_short}_FinStatements_{year}{suffix}.xlsx"
        return str(Path(self.out_dir.get()) / fname)

    def _gen_actual(self) -> None:
        fd = self._get_fd()
        if fd is None:
            return
        out = self._output_path("_Actual")
        try:
            ExcelWriter(fd, proj=None).save(out)
            self._set_status(f"Saved: {out}")
            if messagebox.askyesno("Done", f"Excel saved:\n{out}\n\nOpen it now?"):
                os.startfile(out) if os.name == "nt" else os.system(f'open "{out}"')
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _gen_proj_only(self) -> None:
        fd = self._get_fd()
        if fd is None:
            return
        proj = self._get_proj_inputs()
        if proj is None:
            return
        out = self._output_path("_Projected")
        try:
            ew = ExcelWriter(fd, proj)
            ew.wb.remove   # keep actual sheets too
            ew.save(out)
            self._set_status(f"Saved: {out}")
            if messagebox.askyesno("Done", f"Excel saved:\n{out}\n\nOpen it now?"):
                os.startfile(out) if os.name == "nt" else os.system(f'open "{out}"')
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _gen_all(self) -> None:
        fd = self._get_fd()
        if fd is None:
            return
        proj = self._get_proj_inputs()
        if proj is None:
            return
        out = self._output_path("_Full")
        try:
            ExcelWriter(fd, proj).save(out)
            self._set_status(f"Saved: {out}")
            if messagebox.askyesno("Done", f"Excel saved:\n{out}\n\nOpen it now?"):
                os.startfile(out) if os.name == "nt" else os.system(f'open "{out}"')
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _clear(self) -> None:
        self.db_path.set("")
        self.op_stock.set("")
        self.cl_stock.set("")
        self.fd = None
        self.preview_text.config(state="normal")
        self.preview_text.delete("1.0", "end")
        self.preview_text.config(state="disabled")
        self._set_status("Cleared. Select a Tally SQLite file to begin.")


# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    App()
