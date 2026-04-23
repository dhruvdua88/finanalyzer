"""
FinAnalyzer Pro - Excel Dashboard Builder
Reads from Tally_Source_File_2026-03-12.tsf (SQLite) and creates a multi-sheet Excel dashboard
Green + White theme | Multi-sheet with hyperlink navigation
"""

import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.worksheet.hyperlink import Hyperlink

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = "/Users/dhruvdua88/Documents/Python Testing/Tally Audit App"
TSF_PATH = os.path.join(BASE_DIR, "Tally_Source_File_2026-03-12.tsf")
OUT_PATH = os.path.join(BASE_DIR, "FinAnalyzer_Dashboard.xlsx")

# ── Theme Colors ──────────────────────────────────────────────────────────────
C_DARK_GREEN  = "1B5E20"   # Header bg
C_MED_GREEN   = "2E7D32"   # Sub-header / accent
C_LIGHT_GREEN = "E8F5E9"   # Alternate row
C_MINT        = "C8E6C9"   # KPI card bg
C_WHITE       = "FFFFFF"
C_GOLD        = "F9A825"   # Highlight / warning
C_TEXT_WHITE  = "FFFFFF"
C_TEXT_DARK   = "1A1A1A"
C_BORDER      = "A5D6A7"
C_CHART_GREEN = ["1B5E20","2E7D32","388E3C","43A047","66BB6A","81C784","A5D6A7"]
C_CHART_MULTI = ["1B5E20","F9A825","1565C0","D32F2F","6A1B9A","00838F","558B2F"]

# ── Helpers ───────────────────────────────────────────────────────────────────
def db(sql, path=TSF_PATH):
    con = sqlite3.connect(path)
    cur = con.execute(sql)
    rows = cur.fetchall()
    con.close()
    return rows

def money(v): return f"₹{v:,.0f}"

def mk_font(bold=False, size=11, color=C_TEXT_DARK, name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color)

def mk_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mk_border(color=C_BORDER, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def mk_center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def mk_left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def header_row(ws, row, cols, texts, bg=C_DARK_GREEN, fg=C_TEXT_WHITE, size=11, bold=True):
    for i, (col, text) in enumerate(zip(cols, texts)):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Arial", bold=bold, size=size, color=fg)
        c.fill = mk_fill(bg)
        c.alignment = mk_center(wrap=True)
        c.border = mk_border(color="FFFFFF", style="thin")

def kpi_card(ws, row, col, label, value, sublabel="", bg=C_MINT):
    """Write a KPI card block (3 rows x 2 cols)"""
    # merge label
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    c = ws.cell(row=row, column=col, value=label)
    c.font = Font(name="Arial", bold=True, size=10, color=C_MED_GREEN)
    c.fill = mk_fill(bg)
    c.alignment = mk_center()

    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    v = ws.cell(row=row+1, column=col, value=value)
    v.font = Font(name="Arial", bold=True, size=14, color=C_DARK_GREEN)
    v.fill = mk_fill(bg)
    v.alignment = mk_center()

    if sublabel:
        ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)
        s = ws.cell(row=row+2, column=col, value=sublabel)
        s.font = Font(name="Arial", size=9, color="555555")
        s.fill = mk_fill(bg)
        s.alignment = mk_center()

def nav_button(ws, row, col, label, sheet_name):
    """Fake navigation button using a styled cell with HYPERLINK"""
    c = ws.cell(row=row, column=col, value=f"→ {label}")
    c.font = Font(name="Arial", bold=True, size=10, color=C_TEXT_WHITE)
    c.fill = mk_fill(C_MED_GREEN)
    c.alignment = mk_center()
    c.border = mk_border(color=C_DARK_GREEN, style="medium")
    c.hyperlink = f"#{sheet_name}!A1"

def back_button(ws, row=2, col=1):
    c = ws.cell(row=row, column=col, value="← Back to Summary")
    c.font = Font(name="Arial", bold=True, size=10, color=C_TEXT_WHITE)
    c.fill = mk_fill(C_MED_GREEN)
    c.alignment = mk_center()
    c.border = mk_border(color=C_DARK_GREEN, style="medium")
    c.hyperlink = "#Summary!A1"
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=3)

def sheet_title(ws, title, subtitle="", cols_span=14):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_span)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Arial", bold=True, size=16, color=C_TEXT_WHITE)
    c.fill = mk_fill(C_DARK_GREEN)
    c.alignment = mk_center()
    if subtitle:
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=cols_span)
        s = ws.cell(row=2, column=4, value=subtitle)
        s.font = Font(name="Arial", size=10, color="444444")
        s.alignment = mk_left()

def alt_row(ws, row, col_start, col_end, odd_row):
    color = "F1F8E9" if odd_row else C_WHITE
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = mk_fill(color)

# ══════════════════════════════════════════════════════════════════════════════
# FETCH DATA FROM TSF
# ══════════════════════════════════════════════════════════════════════════════

# 1. Monthly Sales
monthly_sales = db("""
    SELECT substr(date,1,7) as month,
           SUM(ABS(amount)) as sales_amt,
           COUNT(DISTINCT voucher_number) as vouchers
    FROM ledger_entries
    WHERE voucher_type='Sales' AND tally_primary='Sales Accounts'
      AND substr(date,1,7) >= '2025-04' AND substr(date,1,7) <= '2025-10'
    GROUP BY month ORDER BY month
""")

# 2. Monthly P&L
monthly_pnl = db("""
    SELECT substr(date,1,7) as month,
           SUM(CASE WHEN tally_primary='Sales Accounts' THEN ABS(amount) ELSE 0 END) as sales,
           SUM(CASE WHEN tally_primary IN ('Direct Expenses','Indirect Expenses') THEN ABS(amount) ELSE 0 END) as expenses,
           SUM(CASE WHEN tally_primary='Indirect Incomes' THEN ABS(amount) ELSE 0 END) as other_income
    FROM ledger_entries
    WHERE voucher_type NOT IN ('__MASTER_LEDGER__')
      AND substr(date,1,7) >= '2025-04' AND substr(date,1,7) <= '2025-10'
    GROUP BY month ORDER BY month
""")

# 3. Top Customers
top_customers = db("""
    SELECT CASE WHEN instr(party_name,'_') > 0
                THEN substr(party_name,1,instr(party_name,'_')-1)
                ELSE party_name END as customer,
           SUM(ABS(amount)) as total_sales,
           COUNT(DISTINCT voucher_number) as invoices,
           COUNT(DISTINCT substr(date,1,7)) as active_months
    FROM ledger_entries
    WHERE voucher_type='Sales' AND tally_primary='Sundry Debtors'
    GROUP BY customer ORDER BY total_sales DESC LIMIT 10
""")

# 4. Expense Breakdown
expense_heads = db("""
    SELECT tally_parent, SUM(ABS(amount)) as total_exp
    FROM ledger_entries
    WHERE tally_primary IN ('Direct Expenses','Indirect Expenses')
      AND voucher_type != '__MASTER_LEDGER__'
      AND substr(date,1,7) >= '2025-04' AND substr(date,1,7) <= '2025-10'
    GROUP BY tally_parent ORDER BY total_exp DESC LIMIT 10
""")

# 5. Debtors Outstanding
debtors = db("""
    SELECT ledger, closing_balance,
           CASE WHEN closing_balance >= 500000 THEN 'Major'
                WHEN closing_balance >= 100000 THEN 'Medium'
                ELSE 'Minor' END as category
    FROM mst_ledger
    WHERE tally_primary='Sundry Debtors' AND closing_balance > 0
    ORDER BY closing_balance DESC LIMIT 15
""")

# 6. Voucher Type Summary
voucher_summary = db("""
    SELECT voucher_type, COUNT(DISTINCT voucher_number) as count,
           SUM(CASE WHEN amount > 0 THEN amount ELSE 0 END) as debit_total
    FROM ledger_entries
    WHERE voucher_type != '__MASTER_LEDGER__'
    GROUP BY voucher_type ORDER BY count DESC
""")

# 7. Monthly Sales per Top Customer (heatmap data)
customer_monthly = db("""
    SELECT
        CASE WHEN instr(party_name,'_') > 0
             THEN substr(party_name,1,instr(party_name,'_')-1)
             ELSE party_name END as customer,
        substr(date,1,7) as month,
        SUM(ABS(amount)) as amt
    FROM ledger_entries
    WHERE voucher_type='Sales' AND tally_primary='Sundry Debtors'
      AND substr(date,1,7) >= '2025-04' AND substr(date,1,7) <= '2025-10'
    GROUP BY customer, month
    ORDER BY customer, month
""")

# ── Derived KPIs ──────────────────────────────────────────────────────────────
total_sales    = sum(r[1] for r in monthly_sales)
total_expenses = sum(r[2] for r in monthly_pnl)
total_oi       = sum(r[3] for r in monthly_pnl)
net_pnl        = total_sales + total_oi - total_expenses
total_debtors  = sum(r[1] for r in debtors)
months_data    = [r[0] for r in monthly_sales]

# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# Month labels for display
MONTH_LABELS = {
    "2025-04": "Apr-25", "2025-05": "May-25", "2025-06": "Jun-25",
    "2025-07": "Jul-25", "2025-08": "Aug-25", "2025-09": "Sep-25",
    "2025-10": "Oct-25",
}

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws_sum = wb.create_sheet("Summary")
ws_sum.sheet_view.showGridLines = False

# Title banner
ws_sum.merge_cells("A1:N1")
t = ws_sum["A1"]
t.value = "FinAnalyzer Pro  |  Tally Audit Dashboard  |  FY 2025-26"
t.font = Font(name="Arial", bold=True, size=18, color=C_TEXT_WHITE)
t.fill = mk_fill(C_DARK_GREEN)
t.alignment = mk_center()
ws_sum.row_dimensions[1].height = 36

# Sub-header
ws_sum.merge_cells("A2:N2")
s = ws_sum["A2"]
s.value = "Source: Tally_Source_File_2026-03-12.tsf  |  Period: Apr 2025 – Oct 2025  |  Generated by FinAnalyzer Pro"
s.font = Font(name="Arial", size=10, color="444444", italic=True)
s.fill = mk_fill(C_MINT)
s.alignment = mk_center()
ws_sum.row_dimensions[2].height = 20

# ── KPI Cards Row 1 (row 4-6) ─────────────────────────────────────────────
ws_sum.row_dimensions[3].height = 10  # spacer
ws_sum.merge_cells("A4:B4"); ws_sum.merge_cells("A5:B5"); ws_sum.merge_cells("A6:B6")
kpi_card(ws_sum, 4, 1, "Total Sales Revenue", money(total_sales), "Apr-25 to Oct-25")
kpi_card(ws_sum, 4, 4, "Total Expenses", money(total_expenses), "Operating Cost")
kpi_card(ws_sum, 4, 7, "Other Income", money(total_oi), "Non-operating")
kpi_card(ws_sum, 4, 10, "Net P&L", money(net_pnl), "Revenue + OI − Expenses", bg="FFCDD2" if net_pnl < 0 else C_MINT)

for row in [4, 5, 6]:
    ws_sum.row_dimensions[row].height = 22

# ── Navigation Buttons (row 8) ────────────────────────────────────────────
ws_sum.row_dimensions[7].height = 10
ws_sum.merge_cells("A8:C8");  ws_sum.merge_cells("D8:F8")
ws_sum.merge_cells("G8:I8");  ws_sum.merge_cells("J8:L8")
nav_button(ws_sum, 8, 1,  "Sales Register",      "Sales")
nav_button(ws_sum, 8, 4,  "P&L Analysis",        "PL")
nav_button(ws_sum, 8, 7,  "Debtor Ageing",       "Debtors")
nav_button(ws_sum, 8, 10, "Expense Breakdown",   "Expenses")
ws_sum.row_dimensions[8].height = 24

ws_sum.row_dimensions[9].height = 10

# ── Summary: Monthly Sales table (A10 onwards) ───────────────────────────
header_row(ws_sum, 10, [1,2,3,4,5], ["Month","Sales (₹)","Invoices","vs Prev Month","Trend"])
ws_sum.row_dimensions[10].height = 20

for i, (month, sales, inv) in enumerate(monthly_sales):
    r = 11 + i
    ws_sum.row_dimensions[r].height = 18
    odd = i % 2 == 0
    row_bg = "F1F8E9" if odd else C_WHITE
    lbl = MONTH_LABELS.get(month, month)
    for col in range(1, 6):
        ws_sum.cell(row=r, column=col).fill = mk_fill(row_bg)
        ws_sum.cell(row=r, column=col).border = mk_border()
        ws_sum.cell(row=r, column=col).font = Font(name="Arial", size=10)

    ws_sum.cell(row=r, column=1, value=lbl).alignment = mk_center()
    c_sales = ws_sum.cell(row=r, column=2, value=sales)
    c_sales.number_format = '₹#,##0'
    c_sales.alignment = mk_center()
    ws_sum.cell(row=r, column=3, value=inv).alignment = mk_center()

    if i > 0:
        prev_r = r - 1
        chg = ws_sum.cell(row=r, column=4)
        chg.value = f'=B{r}-B{prev_r}'
        chg.number_format = '₹#,##0;(₹#,##0);"-"'
        chg.alignment = mk_center()
        chg.font = Font(name="Arial", size=10)
        trend = ws_sum.cell(row=r, column=5)
        trend.value = f'=IF(D{r}>0,"▲","▼")'
        trend.alignment = mk_center()
    else:
        ws_sum.cell(row=r, column=4, value="-").alignment = mk_center()
        ws_sum.cell(row=r, column=5, value="—").alignment = mk_center()

# Total row
tr = 11 + len(monthly_sales)
ws_sum.row_dimensions[tr].height = 20
header_row(ws_sum, tr, [1,2,3], ["TOTAL", f"=SUM(B11:B{tr-1})", f"=SUM(C11:C{tr-1})"],
           bg=C_MED_GREEN, fg=C_TEXT_WHITE)
ws_sum.cell(row=tr, column=2).number_format = '₹#,##0'

# ── Summary: Voucher Type breakdown (F10) ────────────────────────────────
header_row(ws_sum, 10, [7,8,9], ["Voucher Type","Count","Debit Total (₹)"])
for i, (vtype, cnt, drtotal) in enumerate(voucher_summary[:7]):
    r = 11 + i
    odd = i % 2 == 0
    bg = "F1F8E9" if odd else C_WHITE
    for col in [7,8,9]:
        ws_sum.cell(row=r, column=col).fill = mk_fill(bg)
        ws_sum.cell(row=r, column=col).border = mk_border()
        ws_sum.cell(row=r, column=col).font = Font(name="Arial", size=10)
    ws_sum.cell(row=r, column=7, value=vtype).alignment = mk_left()
    ws_sum.cell(row=r, column=8, value=cnt).alignment = mk_center()
    v = ws_sum.cell(row=r, column=9, value=drtotal)
    v.number_format = '₹#,##0'
    v.alignment = mk_center()

# Column widths Summary
col_widths = {1:12, 2:18, 3:10, 4:18, 5:8, 6:2, 7:32, 8:10, 9:20, 10:28, 11:2, 12:18, 13:2, 14:18}
for col, width in col_widths.items():
    ws_sum.column_dimensions[get_column_letter(col)].width = width

# ── Chart: Monthly Sales bar ──────────────────────────────────────────────
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Monthly Sales — FY 2025-26"
chart1.y_axis.title = "Sales (₹)"
chart1.x_axis.title = "Month"
chart1.style = 10
chart1.width = 18
chart1.height = 12
chart1.grouping = "clustered"

data_ref = Reference(ws_sum, min_col=2, min_row=10, max_row=10+len(monthly_sales))
cats_ref = Reference(ws_sum, min_col=1, min_row=11, max_row=10+len(monthly_sales))
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.series[0].graphicalProperties.solidFill = "2E7D32"
chart1.series[0].graphicalProperties.line.solidFill = "1B5E20"
ws_sum.add_chart(chart1, "A20")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: SALES REGISTER
# ══════════════════════════════════════════════════════════════════════════════
ws_sales = wb.create_sheet("Sales")
ws_sales.sheet_view.showGridLines = False
sheet_title(ws_sales, "Sales Register  |  FinAnalyzer Pro", "FY 2025-26 | Apr – Oct 2025", cols_span=12)
back_button(ws_sales, row=2, col=1)
ws_sales.row_dimensions[1].height = 32
ws_sales.row_dimensions[2].height = 22
ws_sales.row_dimensions[3].height = 10

# Section A: Monthly Sales
header_row(ws_sales, 4, range(1,6), ["Month","Sales Amount (₹)","No. of Invoices","Avg Inv Value (₹)","% of Total"])
ws_sales.row_dimensions[4].height = 20

for i, (month, sales, inv) in enumerate(monthly_sales):
    r = 5 + i
    ws_sales.row_dimensions[r].height = 18
    lbl = MONTH_LABELS.get(month, month)
    odd = i % 2 == 0
    bg = "F1F8E9" if odd else C_WHITE
    for col in range(1, 6):
        ws_sales.cell(row=r, column=col).fill = mk_fill(bg)
        ws_sales.cell(row=r, column=col).border = mk_border()
        ws_sales.cell(row=r, column=col).font = Font(name="Arial", size=10)

    ws_sales.cell(row=r, column=1, value=lbl).alignment = mk_center()
    s_cell = ws_sales.cell(row=r, column=2, value=sales)
    s_cell.number_format = '₹#,##0'
    s_cell.alignment = mk_center()
    ws_sales.cell(row=r, column=3, value=inv).alignment = mk_center()
    avg = ws_sales.cell(row=r, column=4, value=f"=IFERROR(B{r}/C{r},\"-\")")
    avg.number_format = '₹#,##0'
    avg.alignment = mk_center()
    pct = ws_sales.cell(row=r, column=5, value=f"=IFERROR(B{r}/SUM(B5:B{4+len(monthly_sales)}),0)")
    pct.number_format = "0.0%"
    pct.alignment = mk_center()

# Total
tr_s = 5 + len(monthly_sales)
header_row(ws_sales, tr_s, [1,2,3], ["TOTAL", f"=SUM(B5:B{tr_s-1})", f"=SUM(C5:C{tr_s-1})"],
           bg=C_MED_GREEN)
ws_sales.cell(row=tr_s, column=2).number_format = "₹#,##0"
ws_sales.row_dimensions[tr_s].height = 20

# Section B: Top Customers
sec_b_start = tr_s + 2
ws_sales.row_dimensions[tr_s+1].height = 14
ws_sales.merge_cells(start_row=sec_b_start, start_column=1, end_row=sec_b_start, end_column=8)
hdr = ws_sales.cell(row=sec_b_start, column=1, value="Top Customers by Sales Revenue")
hdr.font = Font(name="Arial", bold=True, size=12, color=C_TEXT_WHITE)
hdr.fill = mk_fill(C_DARK_GREEN)
hdr.alignment = mk_left()
ws_sales.row_dimensions[sec_b_start].height = 22

cust_hdr_row = sec_b_start + 1
header_row(ws_sales, cust_hdr_row, range(1,7),
           ["#","Customer Name","Total Sales (₹)","Invoices","Active Months","Avg/Invoice (₹)"])
ws_sales.row_dimensions[cust_hdr_row].height = 20

for i, (cust, sales, inv, months) in enumerate(top_customers):
    r = cust_hdr_row + 1 + i
    ws_sales.row_dimensions[r].height = 18
    odd = i % 2 == 0
    bg = "F1F8E9" if odd else C_WHITE
    for col in range(1, 7):
        ws_sales.cell(row=r, column=col).fill = mk_fill(bg)
        ws_sales.cell(row=r, column=col).border = mk_border()
        ws_sales.cell(row=r, column=col).font = Font(name="Arial", size=10)
    ws_sales.cell(row=r, column=1, value=i+1).alignment = mk_center()
    ws_sales.cell(row=r, column=2, value=cust).alignment = mk_left()
    s2 = ws_sales.cell(row=r, column=3, value=sales)
    s2.number_format = "₹#,##0"; s2.alignment = mk_center()
    ws_sales.cell(row=r, column=4, value=inv).alignment = mk_center()
    ws_sales.cell(row=r, column=5, value=months).alignment = mk_center()
    a2 = ws_sales.cell(row=r, column=6, value=f"=IFERROR(C{r}/D{r},\"-\")")
    a2.number_format = "₹#,##0"; a2.alignment = mk_center()

# ── Customer Monthly Heatmap ──────────────────────────────────────────────
hm_start_row = cust_hdr_row + len(top_customers) + 3
ws_sales.merge_cells(start_row=hm_start_row, start_column=1, end_row=hm_start_row, end_column=10)
hm_hdr = ws_sales.cell(row=hm_start_row, column=1, value="Customer Sales Heatmap (Monthly)")
hm_hdr.font = Font(name="Arial", bold=True, size=12, color=C_TEXT_WHITE)
hm_hdr.fill = mk_fill(C_DARK_GREEN)
hm_hdr.alignment = mk_left()
ws_sales.row_dimensions[hm_start_row].height = 22

months_list = list(MONTH_LABELS.keys())
header_row(ws_sales, hm_start_row+1,
           range(1, 2+len(months_list)),
           ["Customer"] + [MONTH_LABELS[m] for m in months_list],
           bg=C_MED_GREEN)

# Build customer-month dict
cm_dict = {}
for cust, month, amt in customer_monthly:
    cm_dict[(cust, month)] = amt

top_cust_names = [r[0] for r in top_customers[:8]]
for i, cust in enumerate(top_cust_names):
    r = hm_start_row + 2 + i
    ws_sales.row_dimensions[r].height = 18
    odd = i % 2 == 0
    bg = "F1F8E9" if odd else C_WHITE
    ws_sales.cell(row=r, column=1, value=cust[:35]).fill = mk_fill(bg)
    ws_sales.cell(row=r, column=1).border = mk_border()
    ws_sales.cell(row=r, column=1).font = Font(name="Arial", size=9)
    ws_sales.cell(row=r, column=1).alignment = mk_left()
    for j, month in enumerate(months_list):
        col = 2 + j
        val = cm_dict.get((cust, month), 0)
        c = ws_sales.cell(row=r, column=col, value=val)
        c.number_format = '₹#,##0'
        c.alignment = mk_center()
        c.font = Font(name="Arial", size=9)
        # Heat color: deeper green = higher value
        if val > 5000000:
            c.fill = mk_fill("1B5E20"); c.font = Font(name="Arial", size=9, color=C_TEXT_WHITE, bold=True)
        elif val > 1000000:
            c.fill = mk_fill("388E3C"); c.font = Font(name="Arial", size=9, color=C_TEXT_WHITE)
        elif val > 100000:
            c.fill = mk_fill("A5D6A7")
        elif val > 0:
            c.fill = mk_fill("E8F5E9")
        else:
            c.fill = mk_fill(C_WHITE)
        c.border = mk_border()

# Column widths Sales
col_widths_s = {1:4, 2:38, 3:18, 4:12, 5:16, 6:18, 7:2, 8:2}
for col, w in col_widths_s.items():
    ws_sales.column_dimensions[get_column_letter(col)].width = w

# ── Chart: Bar chart top customers ───────────────────────────────────────
chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Top Customers — Total Sales"
chart2.y_axis.title = "Customer"
chart2.x_axis.title = "Sales (₹)"
chart2.style = 10
chart2.width = 20
chart2.height = 14

c2_data_row_start = cust_hdr_row + 1
c2_data_row_end   = cust_hdr_row + len(top_customers)
data2 = Reference(ws_sales, min_col=3, min_row=cust_hdr_row, max_row=c2_data_row_end)
cats2 = Reference(ws_sales, min_col=2, min_row=c2_data_row_start, max_row=c2_data_row_end)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.solidFill = "43A047"
ws_sales.add_chart(chart2, "H5")

# ── Line chart: monthly sales trend ──────────────────────────────────────
chart3 = LineChart()
chart3.title = "Monthly Sales Trend"
chart3.y_axis.title = "₹"
chart3.x_axis.title = "Month"
chart3.style = 10
chart3.width = 18
chart3.height = 10
data3 = Reference(ws_sales, min_col=2, min_row=4, max_row=4+len(monthly_sales))
cats3 = Reference(ws_sales, min_col=1, min_row=5, max_row=4+len(monthly_sales))
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.series[0].graphicalProperties.line.solidFill = "1B5E20"
chart3.series[0].graphicalProperties.line.width = 28000
ws_sales.add_chart(chart3, "H20")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: P&L ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
ws_pl = wb.create_sheet("PL")
ws_pl.sheet_view.showGridLines = False
sheet_title(ws_pl, "Profit & Loss Analysis  |  FinAnalyzer Pro", "FY 2025-26 | Apr – Oct 2025", cols_span=12)
back_button(ws_pl, row=2, col=1)
ws_pl.row_dimensions[1].height = 32
ws_pl.row_dimensions[2].height = 22
ws_pl.row_dimensions[3].height = 10

# Monthly P&L table
header_row(ws_pl, 4, range(1,8),
           ["Month","Sales (₹)","Expenses (₹)","Other Income (₹)","Gross P&L (₹)","Net Margin %","Cost Ratio %"])
ws_pl.row_dimensions[4].height = 22

for i, (month, sales, exp, oi) in enumerate(monthly_pnl[:7]):
    r = 5 + i
    ws_pl.row_dimensions[r].height = 20
    lbl = MONTH_LABELS.get(month, month)
    odd = i % 2 == 0
    bg = "F1F8E9" if odd else C_WHITE
    for col in range(1, 8):
        ws_pl.cell(row=r, column=col).fill = mk_fill(bg)
        ws_pl.cell(row=r, column=col).border = mk_border()
        ws_pl.cell(row=r, column=col).font = Font(name="Arial", size=10)
    ws_pl.cell(row=r, column=1, value=lbl).alignment = mk_center()
    for col, val in [(2, sales), (3, exp), (4, oi)]:
        c = ws_pl.cell(row=r, column=col, value=val)
        c.number_format = "₹#,##0"; c.alignment = mk_center()
    pnl = ws_pl.cell(row=r, column=5, value=f"=B{r}+D{r}-C{r}")
    pnl.number_format = "₹#,##0;(₹#,##0);\"-\""; pnl.alignment = mk_center()
    margin = ws_pl.cell(row=r, column=6, value=f"=IFERROR((B{r}+D{r}-C{r})/B{r},0)")
    margin.number_format = "0.0%"; margin.alignment = mk_center()
    cost_r = ws_pl.cell(row=r, column=7, value=f"=IFERROR(C{r}/B{r},0)")
    cost_r.number_format = "0.0%"; cost_r.alignment = mk_center()

# Totals
tr_pl = 5 + len(monthly_pnl[:7])
header_row(ws_pl, tr_pl, range(1,8),
           ["TOTAL",
            f"=SUM(B5:B{tr_pl-1})", f"=SUM(C5:C{tr_pl-1})", f"=SUM(D5:D{tr_pl-1})",
            f"=B{tr_pl}+D{tr_pl}-C{tr_pl}",
            f"=IFERROR((B{tr_pl}+D{tr_pl}-C{tr_pl})/B{tr_pl},0)",
            f"=IFERROR(C{tr_pl}/B{tr_pl},0)"],
           bg=C_MED_GREEN)
for col, fmt in [(2,"₹#,##0"),(3,"₹#,##0"),(4,"₹#,##0"),(5,"₹#,##0;(₹#,##0);\"-\""),(6,"0.0%"),(7,"0.0%")]:
    ws_pl.cell(row=tr_pl, column=col).number_format = fmt
ws_pl.row_dimensions[tr_pl].height = 22

# Expense Composition Section
ec_start = tr_pl + 2
ws_pl.merge_cells(start_row=ec_start, start_column=1, end_row=ec_start, end_column=6)
ec_hdr = ws_pl.cell(row=ec_start, column=1, value="Expense Composition — Operating Cost Breakdown")
ec_hdr.font = Font(name="Arial", bold=True, size=12, color=C_TEXT_WHITE)
ec_hdr.fill = mk_fill(C_DARK_GREEN)
ec_hdr.alignment = mk_left()
ws_pl.row_dimensions[ec_start].height = 22

header_row(ws_pl, ec_start+1, [1,2,3,4], ["Expense Head","Total Amount (₹)","% of Total Expenses","Cumulative %"])
for i, (head, amt) in enumerate(expense_heads):
    r = ec_start + 2 + i
    ws_pl.row_dimensions[r].height = 18
    odd = i % 2 == 0
    bg = "F1F8E9" if odd else C_WHITE
    for col in range(1, 5):
        ws_pl.cell(row=r, column=col).fill = mk_fill(bg)
        ws_pl.cell(row=r, column=col).border = mk_border()
        ws_pl.cell(row=r, column=col).font = Font(name="Arial", size=10)
    ws_pl.cell(row=r, column=1, value=head).alignment = mk_left()
    a_c = ws_pl.cell(row=r, column=2, value=amt)
    a_c.number_format = "₹#,##0"; a_c.alignment = mk_center()
    pct_c = ws_pl.cell(row=r, column=3, value=f"=IFERROR(B{r}/SUM(B{ec_start+2}:B{ec_start+1+len(expense_heads)}),0)")
    pct_c.number_format = "0.0%"; pct_c.alignment = mk_center()
    if i == 0:
        cum = ws_pl.cell(row=r, column=4, value=f"=C{r}")
    else:
        cum = ws_pl.cell(row=r, column=4, value=f"=D{r-1}+C{r}")
    cum.number_format = "0.0%"; cum.alignment = mk_center()

# Total expenses row
te_row = ec_start + 2 + len(expense_heads)
header_row(ws_pl, te_row, [1,2], ["TOTAL EXPENSES", f"=SUM(B{ec_start+2}:B{te_row-1})"], bg=C_MED_GREEN)
ws_pl.cell(row=te_row, column=2).number_format = "₹#,##0"
ws_pl.row_dimensions[te_row].height = 20

# Column widths P&L
for col, w in {1:12, 2:20, 3:20, 4:18, 5:20, 6:14, 7:14}.items():
    ws_pl.column_dimensions[get_column_letter(col)].width = w

# ── Charts P&L ────────────────────────────────────────────────────────────
# Grouped bar: Sales vs Expenses
chart4 = BarChart()
chart4.type = "col"
chart4.title = "Sales vs Expenses — Monthly"
chart4.y_axis.title = "₹"
chart4.style = 10
chart4.width = 20
chart4.height = 13
chart4.grouping = "clustered"
data4 = Reference(ws_pl, min_col=2, max_col=3, min_row=4, max_row=4+len(monthly_pnl[:7]))
cats4 = Reference(ws_pl, min_col=1, min_row=5, max_row=4+len(monthly_pnl[:7]))
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.series[0].graphicalProperties.solidFill = "2E7D32"
chart4.series[1].graphicalProperties.solidFill = "D32F2F"
ws_pl.add_chart(chart4, "I4")

# Pie chart: Expense composition
chart5 = PieChart()
chart5.title = "Expense Breakdown"
chart5.style = 10
chart5.width = 16
chart5.height = 14
data5 = Reference(ws_pl, min_col=2, min_row=ec_start+1, max_row=ec_start+1+len(expense_heads))
cats5 = Reference(ws_pl, min_col=1, min_row=ec_start+2, max_row=ec_start+1+len(expense_heads))
chart5.add_data(data5, titles_from_data=True)
chart5.set_categories(cats5)
chart5.dataLabels = DataLabelList()
chart5.dataLabels.showPercent = True
chart5.dataLabels.showCatName = False
ws_pl.add_chart(chart5, "I20")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: DEBTORS
# ══════════════════════════════════════════════════════════════════════════════
ws_deb = wb.create_sheet("Debtors")
ws_deb.sheet_view.showGridLines = False
sheet_title(ws_deb, "Debtor Outstanding Analysis  |  FinAnalyzer Pro", "As at 12-Mar-2026", cols_span=10)
back_button(ws_deb, row=2, col=1)
ws_deb.row_dimensions[1].height = 32
ws_deb.row_dimensions[2].height = 22
ws_deb.row_dimensions[3].height = 10

# KPI: Total Debtors
ws_deb.merge_cells("A4:D4")
kpi1 = ws_deb.cell(row=4, column=1, value=f"Total Outstanding Debtors: {money(total_debtors)}")
kpi1.font = Font(name="Arial", bold=True, size=13, color=C_TEXT_WHITE)
kpi1.fill = mk_fill(C_MED_GREEN)
kpi1.alignment = mk_center()
ws_deb.row_dimensions[4].height = 24

ws_deb.merge_cells("E4:H4")
kpi2 = ws_deb.cell(row=4, column=5, value=f"No. of Debtors with Outstanding: {len(debtors)}")
kpi2.font = Font(name="Arial", bold=True, size=13, color=C_TEXT_WHITE)
kpi2.fill = mk_fill(C_MED_GREEN)
kpi2.alignment = mk_center()
ws_deb.row_dimensions[5].height = 12  # spacer

header_row(ws_deb, 6, range(1,7),
           ["#","Debtor / Party Name","Closing Balance (₹)","Category","% of Total","Cumulative %"])
ws_deb.row_dimensions[6].height = 20

total_deb_sum_formula_cells = []
for i, (ledger, bal, cat) in enumerate(debtors):
    r = 7 + i
    ws_deb.row_dimensions[r].height = 18
    odd = i % 2 == 0
    bg = "F1F8E9" if odd else C_WHITE
    for col in range(1, 7):
        ws_deb.cell(row=r, column=col).fill = mk_fill(bg)
        ws_deb.cell(row=r, column=col).border = mk_border()
        ws_deb.cell(row=r, column=col).font = Font(name="Arial", size=10)
    ws_deb.cell(row=r, column=1, value=i+1).alignment = mk_center()
    ws_deb.cell(row=r, column=2, value=ledger).alignment = mk_left()
    b_c = ws_deb.cell(row=r, column=3, value=bal)
    b_c.number_format = "₹#,##0"; b_c.alignment = mk_center()
    # Color by category
    cat_c = ws_deb.cell(row=r, column=4, value=cat)
    cat_c.alignment = mk_center()
    if cat == "Major":
        cat_c.fill = mk_fill("FFCDD2"); cat_c.font = Font(name="Arial", size=10, color="B71C1C", bold=True)
    elif cat == "Medium":
        cat_c.fill = mk_fill("FFF9C4"); cat_c.font = Font(name="Arial", size=10, color="F57F17")
    pct_d = ws_deb.cell(row=r, column=5, value=f"=IFERROR(C{r}/SUM(C7:C{6+len(debtors)}),0)")
    pct_d.number_format = "0.0%"; pct_d.alignment = mk_center()
    if i == 0:
        cum_d = ws_deb.cell(row=r, column=6, value=f"=E{r}")
    else:
        cum_d = ws_deb.cell(row=r, column=6, value=f"=F{r-1}+E{r}")
    cum_d.number_format = "0.0%"; cum_d.alignment = mk_center()

# Total
tr_d = 7 + len(debtors)
header_row(ws_deb, tr_d, [1,2,3], ["TOTAL","",f"=SUM(C7:C{tr_d-1})"], bg=C_MED_GREEN)
ws_deb.cell(row=tr_d, column=3).number_format = "₹#,##0"
ws_deb.row_dimensions[tr_d].height = 22

# Column widths Debtors
for col, w in {1:4, 2:42, 3:20, 4:12, 5:14, 6:14}.items():
    ws_deb.column_dimensions[get_column_letter(col)].width = w

# ── Chart: Debtor bar chart ───────────────────────────────────────────────
chart6 = BarChart()
chart6.type = "bar"
chart6.title = "Top Debtors — Outstanding Balance"
chart6.y_axis.title = "Debtor"
chart6.x_axis.title = "Balance (₹)"
chart6.style = 10
chart6.width = 22
chart6.height = 14
data6 = Reference(ws_deb, min_col=3, min_row=6, max_row=6+len(debtors))
cats6 = Reference(ws_deb, min_col=2, min_row=7, max_row=6+len(debtors))
chart6.add_data(data6, titles_from_data=True)
chart6.set_categories(cats6)
chart6.series[0].graphicalProperties.solidFill = "E53935"
ws_deb.add_chart(chart6, "H6")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: EXPENSES
# ══════════════════════════════════════════════════════════════════════════════
ws_exp = wb.create_sheet("Expenses")
ws_exp.sheet_view.showGridLines = False
sheet_title(ws_exp, "Expense Breakdown Analysis  |  FinAnalyzer Pro", "FY 2025-26 | Apr – Oct 2025", cols_span=10)
back_button(ws_exp, row=2, col=1)
ws_exp.row_dimensions[1].height = 32
ws_exp.row_dimensions[2].height = 22
ws_exp.row_dimensions[3].height = 10

# Full expense table with rank
header_row(ws_exp, 4, range(1, 6),
           ["Rank","Expense Head","Total Amount (₹)","% Share","Running %"])
ws_exp.row_dimensions[4].height = 20

total_exp_amt = sum(r[1] for r in expense_heads)
for i, (head, amt) in enumerate(expense_heads):
    r = 5 + i
    ws_exp.row_dimensions[r].height = 20
    odd = i % 2 == 0
    bg = "F1F8E9" if odd else C_WHITE
    for col in range(1, 6):
        ws_exp.cell(row=r, column=col).fill = mk_fill(bg)
        ws_exp.cell(row=r, column=col).border = mk_border()
        ws_exp.cell(row=r, column=col).font = Font(name="Arial", size=10)
    ws_exp.cell(row=r, column=1, value=i+1).alignment = mk_center()
    ws_exp.cell(row=r, column=2, value=head).alignment = mk_left()
    a = ws_exp.cell(row=r, column=3, value=amt)
    a.number_format = "₹#,##0"; a.alignment = mk_center()
    p = ws_exp.cell(row=r, column=4, value=f"=IFERROR(C{r}/SUM(C5:C{4+len(expense_heads)}),0)")
    p.number_format = "0.0%"; p.alignment = mk_center()
    if i == 0:
        cu = ws_exp.cell(row=r, column=5, value=f"=D{r}")
    else:
        cu = ws_exp.cell(row=r, column=5, value=f"=E{r-1}+D{r}")
    cu.number_format = "0.0%"; cu.alignment = mk_center()

tr_e = 5 + len(expense_heads)
header_row(ws_exp, tr_e, [1,2,3], ["","TOTAL",f"=SUM(C5:C{tr_e-1})"], bg=C_MED_GREEN)
ws_exp.cell(row=tr_e, column=3).number_format = "₹#,##0"
ws_exp.row_dimensions[tr_e].height = 22

# Column widths Expenses
for col, w in {1:6, 2:38, 3:20, 4:12, 5:14}.items():
    ws_exp.column_dimensions[get_column_letter(col)].width = w

# ── Charts Expenses ───────────────────────────────────────────────────────
chart7 = BarChart()
chart7.type = "bar"
chart7.title = "Expense Heads — Total Spend"
chart7.style = 10
chart7.width = 22
chart7.height = 14
data7 = Reference(ws_exp, min_col=3, min_row=4, max_row=4+len(expense_heads))
cats7 = Reference(ws_exp, min_col=2, min_row=5, max_row=4+len(expense_heads))
chart7.add_data(data7, titles_from_data=True)
chart7.set_categories(cats7)
chart7.series[0].graphicalProperties.solidFill = "1565C0"
ws_exp.add_chart(chart7, "G4")

chart8 = PieChart()
chart8.title = "Expense Mix"
chart8.style = 10
chart8.width = 18
chart8.height = 14
data8 = Reference(ws_exp, min_col=3, min_row=4, max_row=4+len(expense_heads))
cats8 = Reference(ws_exp, min_col=2, min_row=5, max_row=4+len(expense_heads))
chart8.add_data(data8, titles_from_data=True)
chart8.set_categories(cats8)
chart8.dataLabels = DataLabelList()
chart8.dataLabels.showPercent = True
ws_exp.add_chart(chart8, "G20")


# ══════════════════════════════════════════════════════════════════════════════
# FINAL SAVE
# ══════════════════════════════════════════════════════════════════════════════
wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
