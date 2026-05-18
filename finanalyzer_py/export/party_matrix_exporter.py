from __future__ import annotations

from openpyxl.utils import get_column_letter
from export.base_exporter import BaseExporter
from export.styles import (
    AMOUNT_FMT, COUNT_FMT, hex_fill, HEADER_FONT, HEADER_FILL, CENTER,
    thin_border, TOTAL_FONT, TOTAL_FILL, medium_border,
    SALES_COL, PURCH_COL, EXP_COL, TDS_COL, GST_COL, RCM_COL, BANK_COL, OTHER_COL,
    RED_FILL, AMBER_FILL,
)

SUMMARY_HEADERS = [
    "Party / Ledger Name", "Voucher Count", "First Date", "Last Date",
    "Total Sales (₹)", "Total Purchase (₹)", "Total Expenses (₹)",
    "TDS Deducted (₹)", "TDS / Expense %", "GST (₹)", "GST / (Sales+Exp) %",
    "RCM (₹)", "Bank (₹)", "Others / Adj (₹)", "Net Balance (+Cr / -Dr)",
    "Top Expense / Purchase Ledgers",
]
VOUCHER_HEADERS = [
    "Party", "Date", "Voucher Type", "Voucher No",
    "Party Amount (+Cr/-Dr)", "Expense (₹)", "Sales (₹)", "Purchase (₹)",
    "TDS (₹)", "GST (₹)", "RCM (₹)", "Bank (₹)", "Others (₹)",
    "Counter-Ledger Breakdown",
]
ANOMALY_HEADERS = ["Anomaly Type", "Party", "Metric", "Value (₹)", "Note"]


class PartyMatrixExporter(BaseExporter):

    def export(self, data: dict, company: str, period: str) -> None:
        meta = f"{company}  |  {period}"
        self._summary_sheet(data["parties"], meta)
        self._voucher_sheet(data["vouchers"], meta)
        self._anomaly_sheet(data["anomalies"], meta)
        self._tagged_ledger_sheet(data["tagged_ledgers"], meta)

    def _summary_sheet(self, parties: list[dict], meta: str) -> None:
        ws = self.add_sheet("Summary")
        n = len(SUMMARY_HEADERS)
        self.write_title_row(ws, "Party Ledger Matrix – Summary", n, row=1)
        self.write_meta_row(ws, meta, n, row=2)
        self.write_header_row(ws, SUMMARY_HEADERS, row=4, freeze_row=4, freeze_col=1)

        amt_cols = {i: AMOUNT_FMT for i in [5, 6, 7, 8, 10, 12, 13, 14, 15]}
        amt_cols[2] = COUNT_FMT
        pct_cols = {9: "0.00", 11: "0.00"}

        for i, p in enumerate(parties):
            row_fill = None
            if p["expense"] > 0.01 and p["tds"] < 0.01:
                row_fill = RED_FILL
            elif (p["sales"] + p["expense"]) > 0.01 and p["gst"] < 0.01:
                row_fill = AMBER_FILL

            vals = [
                p["party"], p["voucher_count"], p["first_date"], p["last_date"],
                p["sales"], p["purchase"], p["expense"],
                p["tds"], p["tds_pct"], p["gst"], p["gst_pct"],
                p["rcm"], p["bank"], p["others"], p["net_balance"],
                p["top_ledgers"],
            ]
            num_fmts = {**amt_cols, **pct_cols}
            self.write_data_row(ws, i + 5, vals, i % 2 == 0, num_fmt_cols=num_fmts,
                                fill_override=row_fill)

        self.set_col_widths(ws, [34, 10, 12, 12, 14, 14, 14, 14, 14, 14, 16, 12, 12, 14, 18, 60])

    def _voucher_sheet(self, vouchers: list[dict], meta: str) -> None:
        ws = self.add_sheet("Voucher Detail")
        n = len(VOUCHER_HEADERS)
        self.write_title_row(ws, "Voucher Detail", n, row=1)
        self.write_meta_row(ws, meta, n, row=2)
        self.write_header_row(ws, VOUCHER_HEADERS, row=4, freeze_row=4, freeze_col=1)

        amt_cols = {i: AMOUNT_FMT for i in range(5, 14)}
        for i, v in enumerate(vouchers):
            vals = [
                v["party"], v["date"], v["voucher_type"], v["voucher_number"],
                v["party_amount"], v["expense"], v["sales"], v["purchase"],
                v["tds"], v["gst"], v["rcm"], v["bank"], v["others"],
                v["counter_breakdown"],
            ]
            self.write_data_row(ws, i + 5, vals, i % 2 == 0, num_fmt_cols=amt_cols)

        ws.freeze_panes = ws.cell(row=5, column=2)
        self.set_col_widths(ws, [30, 11, 18, 14, 18, 14, 12, 14, 12, 12, 12, 12, 12, 80])

    def _anomaly_sheet(self, anomalies: list[dict], meta: str) -> None:
        ws = self.add_sheet("Anomalies")
        n = len(ANOMALY_HEADERS)
        self.write_title_row(ws, "Anomalies", n, row=1)
        self.write_meta_row(ws, meta, n, row=2)
        self.write_header_row(ws, ANOMALY_HEADERS, row=4, freeze_row=4)

        ANOM_COLORS = {
            "Zero TDS": "FEE2E2",
            "Zero GST": "FEF3C7",
            "Balance Gap": "FFEDD5",
            "High Others": "F1F5F9",
        }
        for i, a in enumerate(anomalies):
            fill = hex_fill(ANOM_COLORS.get(a["anomaly_type"], "FFFFFF"))
            vals = [a["anomaly_type"], a["party"], a["metric"], a["value"], a["note"]]
            self.write_data_row(ws, i + 5, vals, True, num_fmt_cols={4: AMOUNT_FMT}, fill_override=fill)

        self.set_col_widths(ws, [14, 34, 20, 16, 70])

    def _tagged_ledger_sheet(self, tagged: dict, meta: str) -> None:
        ws = self.add_sheet("Tagged Ledgers")
        self.write_title_row(ws, "Tagged Ledgers Configuration", 3, row=1)
        self.write_meta_row(ws, meta, 3, row=2)

        # Headers
        for col, (label, color) in enumerate([("TDS Ledgers", TDS_COL), ("GST Ledgers", GST_COL), ("RCM Ledgers", RCM_COL)], 1):
            c = ws.cell(row=3, column=col, value=label)
            c.font = HEADER_FONT
            c.fill = hex_fill(color)
            c.alignment = CENTER
            c.border = thin_border()

        max_rows = max(len(tagged["tds"]), len(tagged["gst"]), len(tagged["rcm"]))
        for r in range(max_rows):
            for col, key in enumerate(["tds", "gst", "rcm"], 1):
                items = tagged[key]
                val = items[r] if r < len(items) else ""
                c = ws.cell(row=r + 4, column=col, value=val)
                c.border = thin_border()

        self.set_col_widths(ws, [40, 40, 40])
