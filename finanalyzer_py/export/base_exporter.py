from __future__ import annotations

from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from export.styles import (
    TITLE_FONT, TITLE_FILL, META_FONT, META_FILL,
    HEADER_FONT, HEADER_FILL, DATA_FONT, TOTAL_FONT, TOTAL_FILL,
    OBS_FONT, OBS_FILL, ODD_FILL, EVEN_FILL, BOLD_FONT,
    CENTER, LEFT, RIGHT, WRAP_LEFT,
    AMOUNT_FMT, COUNT_FMT, DATE_FMT,
    thin_border, medium_border, top_medium_border,
)


class BaseExporter:
    def __init__(self) -> None:
        self.wb = Workbook()
        # Remove the default blank sheet
        self.wb.remove(self.wb.active)

    def add_sheet(self, title: str) -> Worksheet:
        return self.wb.create_sheet(title=title)

    # ── Title / metadata rows ─────────────────────────────────────────────────

    def write_title_row(
        self, ws: Worksheet, title: str, col_count: int, row: int = 1
    ) -> None:
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=col_count
        )
        c = ws.cell(row=row, column=1, value=title)
        c.font = TITLE_FONT
        c.fill = TITLE_FILL
        c.alignment = CENTER
        ws.row_dimensions[row].height = 28

    def write_meta_row(
        self,
        ws: Worksheet,
        meta_text: str,
        col_count: int,
        row: int = 2,
    ) -> None:
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=col_count
        )
        c = ws.cell(row=row, column=1, value=meta_text)
        c.font = META_FONT
        c.fill = META_FILL
        c.alignment = LEFT

    def write_header_row(
        self,
        ws: Worksheet,
        headers: list[str],
        row: int,
        freeze_col: int = 0,
        freeze_row: int = 0,
    ) -> None:
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.border = thin_border()
            c.alignment = CENTER
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A{row}:{last_col}{row}"
        if freeze_col or freeze_row:
            fc = freeze_col + 1 if freeze_col else 1
            fr = freeze_row + 1 if freeze_row else 1
            ws.freeze_panes = ws.cell(row=fr, column=fc)

    def write_data_row(
        self,
        ws: Worksheet,
        row: int,
        values: list,
        is_odd: bool = True,
        num_fmt_cols: dict[int, str] | None = None,
        align_cols: dict[int, "Alignment"] | None = None,
        fill_override: "PatternFill | None" = None,
    ) -> None:
        fill = fill_override if fill_override else (ODD_FILL if is_odd else EVEN_FILL)
        num_fmt_cols = num_fmt_cols or {}
        align_cols = align_cols or {}
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = DATA_FONT
            c.fill = fill
            c.border = thin_border()
            if col in align_cols:
                c.alignment = align_cols[col]
            elif isinstance(val, (int, float)) and not isinstance(val, bool):
                c.alignment = RIGHT
            else:
                c.alignment = LEFT
            if col in num_fmt_cols:
                c.number_format = num_fmt_cols[col]
            elif isinstance(val, float):
                c.number_format = AMOUNT_FMT
            elif isinstance(val, int) and not isinstance(val, bool):
                c.number_format = COUNT_FMT

    def write_total_row(
        self,
        ws: Worksheet,
        row: int,
        values: list,
        num_fmt_cols: dict[int, str] | None = None,
    ) -> None:
        num_fmt_cols = num_fmt_cols or {}
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = TOTAL_FONT
            c.fill = TOTAL_FILL
            c.border = medium_border()
            c.alignment = RIGHT if isinstance(val, (int, float)) and not isinstance(val, bool) else LEFT
            if col in num_fmt_cols:
                c.number_format = num_fmt_cols[col]
            elif isinstance(val, float):
                c.number_format = AMOUNT_FMT

    def write_obs_block(
        self, ws: Worksheet, start_row: int, observations: list[str], col_count: int
    ) -> None:
        for i, obs in enumerate(observations):
            r = start_row + i
            ws.merge_cells(
                start_row=r, start_column=1, end_row=r, end_column=col_count
            )
            c = ws.cell(row=r, column=1, value=obs)
            c.font = OBS_FONT
            c.fill = OBS_FILL
            c.alignment = WRAP_LEFT
            c.border = thin_border()

    def write_super_header(
        self,
        ws: Worksheet,
        row: int,
        blocks: list[tuple[int, int, str, str]],  # (start_col, end_col, label, color_hex)
    ) -> None:
        from export.styles import hex_fill
        for start_col, end_col, label, color in blocks:
            ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col
            )
            c = ws.cell(row=row, column=start_col, value=label)
            c.font = HEADER_FONT
            c.fill = hex_fill(color)
            c.alignment = CENTER
            c.border = thin_border()

    def set_col_widths(self, ws: Worksheet, widths: list[int | float]) -> None:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def autofit_columns(
        self, ws: Worksheet, min_width: float = 8, max_width: float = 40
    ) -> None:
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in col), default=0
            )
            letter = get_column_letter(col[0].column)
            ws.column_dimensions[letter].width = min(
                max(max_len + 2, min_width), max_width
            )

    def save(self, path: str) -> None:
        self.wb.save(path)

    @staticmethod
    def today_str() -> str:
        return date.today().strftime("%Y-%m-%d")
