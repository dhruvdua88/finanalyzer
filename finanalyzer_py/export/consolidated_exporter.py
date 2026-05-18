"""
Consolidated Excel export — all analysis modules in one timestamped workbook.

Sheet order:
  00_Cover              – summary KPIs + hyperlinks to every sheet
  01_Trial_Balance_Grp  – Trial Balance (Group Summary)
  02_Trial_Balance_Ledg – Trial Balance (Ledger Detail)
  03_Ledger_Analytics   – Ledger Analytics
  04_BS_Cleanliness     – Balance Sheet Cleanliness
  05_Debtor_Ageing      – Debtor Ageing Summary
  06_Debtor_FIFO        – Debtor Invoice FIFO
  07_Debtor_Recon       – Debtor Reconciliation
  08_Creditor_Ageing    – Creditor Ageing Summary
  09_Creditor_FIFO      – Creditor Invoice FIFO
  10_Creditor_Recon     – Creditor Reconciliation
  11_Voucher_Book       – Voucher Book
  12_Party_Matrix       – Party Ledger Matrix Summary
  13_Party_Vouchers     – Party Matrix Voucher Detail
  14_Party_Anomalies    – Party Matrix Anomalies
  15_Party_Tags         – Tagged Ledgers
  16_Cash_Flow          – Cash Flow
  17_PnL                – Profit & Loss
  18_GST_Rate           – GST Rate Analysis
  19_Sales_Register     – Sales Register
  20_Purchase_Register  – Purchase GST Register
  21_GST_Ledger         – GST Ledger Summary
  22_RCM                – RCM Analysis
  23_Blocked_Credit     – GST Blocked Credit
  24_TDS_Analysis       – TDS Analysis (20 cols)
  25_GSTR2B_Summary     – GSTR-2B Position Summary
  26_GSTR2B_Detail      – GSTR-2B Reconciliation Detail
  27_Exception_Heatmap  – Exception Density Heatmap
  28_Orphan_PL          – Orphan P&L Vouchers
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from export.styles import (
    NAVY, TEAL, WHITE, LGRAY, LBLUE,
    TITLE_FONT, TITLE_FILL, META_FONT, META_FILL, HEADER_FONT, HEADER_FILL,
    TOTAL_FONT, TOTAL_FILL, ODD_FILL, EVEN_FILL,
    CENTER, LEFT, RIGHT, thin_border, medium_border,
    AMOUNT_FMT, COUNT_FMT,
)
from app.state import AppState


def export_consolidated(
    state: AppState,
    output_dir: str | Path = ".",
    on_progress: Callable[[str], None] | None = None,
) -> Path:
    """
    Run all analysis modules and write one consolidated Excel workbook.
    Returns the path of the saved file.
    Called from a background thread — do NOT touch tkinter widgets here.
    """
    def progress(msg: str) -> None:
        if on_progress:
            on_progress(msg)

    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    company = state.settings.company_name or "Company"
    fy = state.settings.fiscal_year or ""
    filename = f"FinAnalyzer_Report_{ts}.xlsx"
    output_path = Path(output_dir) / filename

    wb = Workbook()
    wb.remove(wb.active)

    entries = state.entries
    master = state.master_entries
    settings = state.settings
    meta = f"{company}  |  {fy}  |  Generated {datetime.now().strftime('%d %b %Y %H:%M')}"

    sheet_index: list[tuple[str, str]] = []  # (sheet_name, description)

    # ── Cover sheet (placeholder — filled at the end with real links) ─────────
    progress("Building cover sheet…")
    cover = wb.create_sheet("00_Cover")

    # ── Trial Balance ─────────────────────────────────────────────────────────
    progress("Trial Balance…")
    from analysis.trial_balance import compute_trial_balance
    from export.trial_balance_exporter import TrialBalanceExporter
    tb_data = compute_trial_balance(entries, master)
    tb_exp = TrialBalanceExporter(wb=wb)
    tb_exp._export_groups(tb_data["groups"], company, fy)
    tb_exp._export_ledgers(tb_data["ledgers"], company, fy)
    # Rename to indexed names
    _rename_last(wb, "01_Trial_Bal_Groups", "02_Trial_Bal_Ledgers")
    sheet_index += [("01_Trial_Bal_Groups", "Trial Balance – Group Summary"),
                    ("02_Trial_Bal_Ledgers", "Trial Balance – Ledger Detail")]

    # ── Ledger Analytics ──────────────────────────────────────────────────────
    progress("Ledger Analytics…")
    from analysis.ledger_analytics import compute_ledger_analytics
    from export.misc_exporters import LedgerAnalyticsExporter
    la_data = compute_ledger_analytics(entries, master)
    LedgerAnalyticsExporter(wb=wb).export(la_data, company, fy)
    _rename_last(wb, "03_Ledger_Analytics")
    sheet_index.append(("03_Ledger_Analytics", "Ledger Analytics"))

    # ── BS Cleanliness ────────────────────────────────────────────────────────
    progress("BS Cleanliness…")
    from analysis.misc_analysis import compute_bs_cleanliness
    from export.misc_exporters import BSCleanlinessExporter
    bs_data = compute_bs_cleanliness(entries, master)
    BSCleanlinessExporter(wb=wb).export(bs_data, company, fy)
    _rename_last(wb, "04_BS_Cleanliness")
    sheet_index.append(("04_BS_Cleanliness", "Balance Sheet Cleanliness"))

    # ── Debtor Ageing ─────────────────────────────────────────────────────────
    progress("Debtor Ageing (FIFO)…")
    from analysis.ageing import compute_ageing
    from export.ageing_exporter import AgeingExporter
    debtor_data = compute_ageing(entries, mode="debtor",
                                 as_of_str=settings.as_of_date)
    d_exp = AgeingExporter(wb=wb)
    d_exp._summary_sheet(debtor_data["parties"], "Debtor", meta)
    d_exp._invoice_sheet(debtor_data["parties"], "Debtor", meta)
    d_exp._recon_sheet(debtor_data["recon"], "Debtor", meta)
    _rename_last(wb, "05_Debtor_Ageing", "06_Debtor_FIFO", "07_Debtor_Recon")
    sheet_index += [("05_Debtor_Ageing", "Debtor Ageing Summary"),
                    ("06_Debtor_FIFO", "Debtor Invoice FIFO Detail"),
                    ("07_Debtor_Recon", "Debtor FIFO Reconciliation")]

    # ── Creditor Ageing ───────────────────────────────────────────────────────
    progress("Creditor Ageing (FIFO)…")
    creditor_data = compute_ageing(entries, mode="creditor",
                                   as_of_str=settings.as_of_date)
    c_exp = AgeingExporter(wb=wb)
    c_exp._summary_sheet(creditor_data["parties"], "Creditor", meta)
    c_exp._invoice_sheet(creditor_data["parties"], "Creditor", meta)
    c_exp._recon_sheet(creditor_data["recon"], "Creditor", meta)
    _rename_last(wb, "08_Creditor_Ageing", "09_Creditor_FIFO", "10_Creditor_Recon")
    sheet_index += [("08_Creditor_Ageing", "Creditor Ageing Summary"),
                    ("09_Creditor_FIFO", "Creditor Invoice FIFO Detail"),
                    ("10_Creditor_Recon", "Creditor FIFO Reconciliation")]

    # ── Voucher Book ──────────────────────────────────────────────────────────
    progress("Voucher Book…")
    from analysis.voucher_book import compute_voucher_book
    from export.misc_exporters import VoucherBookExporter
    vb_data = compute_voucher_book(entries)
    VoucherBookExporter(wb=wb).export(vb_data, company, fy)
    _rename_last(wb, "11_Voucher_Book")
    sheet_index.append(("11_Voucher_Book", "Voucher Book"))

    # ── Party Ledger Matrix ───────────────────────────────────────────────────
    progress("Party Ledger Matrix…")
    from analysis.party_matrix import compute_party_matrix
    from export.party_matrix_exporter import PartyMatrixExporter
    pm_data = compute_party_matrix(entries, settings)
    pm_exp = PartyMatrixExporter(wb=wb)
    pm_exp._summary_sheet(pm_data["parties"], meta)
    pm_exp._voucher_sheet(pm_data["vouchers"], meta)
    pm_exp._anomaly_sheet(pm_data["anomalies"], meta)
    pm_exp._tagged_ledger_sheet(pm_data["tagged_ledgers"], meta)
    _rename_last(wb, "12_Party_Matrix", "13_Party_Vouchers",
                 "14_Party_Anomalies", "15_Party_Tags")
    sheet_index += [("12_Party_Matrix", "Party Ledger Matrix – Summary"),
                    ("13_Party_Vouchers", "Party Matrix – Voucher Detail"),
                    ("14_Party_Anomalies", "Party Matrix – Anomalies"),
                    ("15_Party_Tags", "Tagged Ledgers Configuration")]

    # ── Cash Flow ─────────────────────────────────────────────────────────────
    progress("Cash Flow…")
    from analysis.cash_flow import compute_cash_flow
    from export.misc_exporters import CashFlowExporter
    cf_data = compute_cash_flow(entries, master)
    CashFlowExporter(wb=wb).export(cf_data, company, fy)
    _rename_last(wb, "16_Cash_Flow")
    sheet_index.append(("16_Cash_Flow", "Cash Flow Analysis"))

    # ── P&L Analysis ──────────────────────────────────────────────────────────
    progress("P&L Analysis…")
    from analysis.pnl_analysis import compute_pnl
    from export.misc_exporters import PnLExporter
    pnl_data = compute_pnl(entries)
    PnLExporter(wb=wb).export(pnl_data, company, fy)
    _rename_last(wb, "17_PnL_Analysis")
    sheet_index.append(("17_PnL_Analysis", "Profit & Loss Analysis"))

    # ── GST Rate Analysis ─────────────────────────────────────────────────────
    progress("GST Rate Analysis…")
    from analysis.gst_analysis import (
        compute_gst_rate_analysis, compute_sales_register,
        compute_purchase_register, compute_gst_ledger_summary,
        compute_rcm_analysis, compute_blocked_credit,
    )
    from export.misc_exporters import (
        GSTRateExporter, SalesRegisterExporter, PurchaseRegisterExporter,
        GSTLedgerExporter, RCMExporter, GSTExpenseExporter,
    )
    GSTRateExporter(wb=wb).export(compute_gst_rate_analysis(entries, settings), company, fy)
    _rename_last(wb, "18_GST_Rate")
    sheet_index.append(("18_GST_Rate", "GST Rate Analysis"))

    progress("Sales Register…")
    SalesRegisterExporter(wb=wb).export(compute_sales_register(entries, settings), company, fy)
    _rename_last(wb, "19_Sales_Register")
    sheet_index.append(("19_Sales_Register", "Sales Register"))

    progress("Purchase GST Register…")
    PurchaseRegisterExporter(wb=wb).export(compute_purchase_register(entries, settings), company, fy)
    _rename_last(wb, "20_Purchase_Register")
    sheet_index.append(("20_Purchase_Register", "Purchase GST Register"))

    progress("GST Ledger Summary…")
    GSTLedgerExporter(wb=wb).export(compute_gst_ledger_summary(entries, settings), company, fy)
    _rename_last(wb, "21_GST_Ledger")
    sheet_index.append(("21_GST_Ledger", "GST Ledger Summary"))

    progress("RCM Analysis…")
    RCMExporter(wb=wb).export(compute_rcm_analysis(entries, settings), company, fy)
    _rename_last(wb, "22_RCM_Analysis")
    sheet_index.append(("22_RCM_Analysis", "RCM Analysis"))

    progress("GST Blocked Credit…")
    GSTExpenseExporter(wb=wb).export(compute_blocked_credit(entries, settings), company, fy)
    _rename_last(wb, "23_Blocked_Credit")
    sheet_index.append(("23_Blocked_Credit", "GST Expense – Blocked Credit"))

    # ── TDS Analysis ──────────────────────────────────────────────────────────
    progress("TDS Analysis…")
    from analysis.tds_analysis import compute_tds
    from export.tds_exporter import TDSExporter
    TDSExporter(wb=wb).export(compute_tds(entries, settings), company, fy)
    _rename_last(wb, "24_TDS_Analysis")
    sheet_index.append(("24_TDS_Analysis", "TDS Analysis (20 columns)"))

    # ── GSTR-2B Reconciliation ────────────────────────────────────────────────
    progress("GSTR-2B Reconciliation…")
    from analysis.misc_analysis import compute_gstr2b_reco
    from export.misc_exporters import GSTR2BExporter
    g2b_data = compute_gstr2b_reco(entries, state.gstr2b_b2b_rows, settings)
    g2b_exp = GSTR2BExporter(wb=wb)
    g2b_exp._summary_sheet(g2b_data["summary"], meta)
    g2b_exp._detail_sheet(g2b_data["rows"], meta)
    _rename_last(wb, "25_GSTR2B_Summary", "26_GSTR2B_Detail")
    sheet_index += [("25_GSTR2B_Summary", "GSTR-2B Position Summary"),
                    ("26_GSTR2B_Detail", "GSTR-2B Reconciliation Detail")]

    # ── Exception Heatmap ─────────────────────────────────────────────────────
    progress("Exception Heatmap…")
    from analysis.exception_heatmap import compute_exception_heatmap
    from export.misc_exporters import ExceptionHeatmapExporter
    ExceptionHeatmapExporter(wb=wb).export(compute_exception_heatmap(entries, settings), company, fy)
    _rename_last(wb, "27_Exception_Heatmap")
    sheet_index.append(("27_Exception_Heatmap", "Exception Density Heatmap"))

    # ── Orphan P&L ────────────────────────────────────────────────────────────
    progress("Orphan P&L Vouchers…")
    from analysis.misc_analysis import compute_orphan_pl
    from export.misc_exporters import OrphanPLExporter
    OrphanPLExporter(wb=wb).export(compute_orphan_pl(entries), company, fy)
    _rename_last(wb, "28_Orphan_PL")
    sheet_index.append(("28_Orphan_PL", "Orphan P&L Vouchers (No BS Offset)"))

    # ── Related Party (AS-18) ─────────────────────────────────────────────────
    progress("Related Party Analysis…")
    from analysis.misc_analysis import compute_related_party
    from export.misc_exporters import RelatedPartyExporter
    rp_data = compute_related_party(entries, settings)
    rp_exp = RelatedPartyExporter(wb=wb)
    rp_exp._build_parties_sheet(rp_data["parties"], meta, wb)
    rp_exp._build_txn_sheet(rp_data["transactions"], meta, wb)
    _rename_last(wb, "29_Related_Parties", "30_Related_Txns")
    sheet_index += [("29_Related_Parties", "Related Party Summary (AS-18)"),
                    ("30_Related_Txns", "Related Party Transactions")]

    # ── Variance Analysis ─────────────────────────────────────────────────────
    progress("Variance Analysis…")
    from analysis.misc_analysis import compute_variance
    from export.misc_exporters import VarianceExporter
    var_rows = compute_variance(entries)
    months_for_var: list[str] = []
    if var_rows:
        months_for_var = [k for k in var_rows[0].keys() if k != "ledger" and "_pct" not in k]
    VarianceExporter(wb=wb).export(var_rows, months_for_var, company, fy)
    _rename_last(wb, "31_Variance")
    sheet_index.append(("31_Variance", "Month-on-Month Variance Analysis"))

    # ── Fill in the Cover sheet ───────────────────────────────────────────────
    progress("Writing cover sheet…")
    _build_cover(cover, wb, company, fy, meta, sheet_index, state)

    # Move cover to first position
    wb.move_sheet("00_Cover", offset=-len(wb.sheetnames))

    progress(f"Saving {filename}…")
    wb.save(str(output_path))
    return output_path


# ── Helpers ───────────────────────────────────────────────────────────────────

def _rename_last(wb: Workbook, *new_names: str) -> None:
    """Rename the last N sheets added to the workbook."""
    sheets = wb.sheetnames
    for i, name in enumerate(reversed(new_names), 1):
        old = sheets[-i]
        if old != name:
            wb[old].title = name


def _build_cover(
    ws,
    wb: Workbook,
    company: str,
    fy: str,
    meta: str,
    sheet_index: list[tuple[str, str]],
    state: AppState,
) -> None:
    """Build a rich cover / index sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "FinAnalyzer – Consolidated Audit Report"
    c.font = Font(bold=True, size=22, color=WHITE, name="Calibri")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:D2")
    c2 = ws["A2"]
    c2.value = meta
    c2.font = Font(bold=True, size=11, color=NAVY, name="Calibri")
    c2.fill = PatternFill("solid", fgColor=LBLUE)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── KPI summary block ─────────────────────────────────────────────────────
    kpis = [
        ("Total Journal Entries", f"{state.total_rows:,}"),
        ("Unique Vouchers", f"{state.unique_vouchers:,}"),
        ("Date Range", f"{state.min_date}  →  {state.max_date}"),
        ("Fiscal Months", str(len(state.available_months))),
        ("Distinct Ledgers", str(len(state.all_ledger_names))),
        ("GSTR-2B Invoices", f"{len(state.gstr2b_b2b_rows):,}"),
    ]
    ws.row_dimensions[4].height = 16
    ws["A4"].value = "DATASET SUMMARY"
    ws["A4"].font = Font(bold=True, size=10, color=NAVY, name="Calibri")

    for i, (label, value) in enumerate(kpis):
        row = 5 + i
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(size=10, name="Calibri")
        lc.fill = PatternFill("solid", fgColor=LGRAY)
        lc.border = _thin()
        lc.alignment = Alignment(horizontal="left", vertical="center")

        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(bold=True, size=10, name="Calibri")
        vc.border = _thin()
        vc.alignment = Alignment(horizontal="right", vertical="center")

    # ── Sheet index table ─────────────────────────────────────────────────────
    idx_start = 5 + len(kpis) + 2

    # Header
    ws.merge_cells(start_row=idx_start, start_column=1, end_row=idx_start, end_column=4)
    hc = ws.cell(row=idx_start, column=1, value="CONTENTS — Click a sheet name to navigate")
    hc.font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    hc.fill = PatternFill("solid", fgColor=NAVY)
    hc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[idx_start].height = 20

    col_heads = ["#", "Sheet Name", "Description", "Rows"]
    for col, h in enumerate(col_heads, 1):
        c = ws.cell(row=idx_start + 1, column=col, value=h)
        c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.fill = PatternFill("solid", fgColor="374151")
        c.border = _thin()
        c.alignment = Alignment(horizontal="center")

    teal_fill = PatternFill("solid", fgColor=TEAL)
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    even_fill = PatternFill("solid", fgColor=LGRAY)

    for i, (sheet_name, description) in enumerate(sheet_index):
        row = idx_start + 2 + i
        fill = odd_fill if i % 2 == 0 else even_fill
        ws.row_dimensions[row].height = 16

        # Number
        nc = ws.cell(row=row, column=1, value=i + 1)
        nc.font = Font(size=9, color=WHITE, name="Calibri")
        nc.fill = teal_fill
        nc.border = _thin()
        nc.alignment = Alignment(horizontal="center")

        # Sheet link
        lc = ws.cell(row=row, column=2, value=sheet_name)
        if sheet_name in wb.sheetnames:
            lc.hyperlink = f"#{sheet_name}!A1"
        lc.font = Font(size=10, color="1D4ED8", underline="single", name="Calibri")
        lc.fill = fill
        lc.border = _thin()
        lc.alignment = Alignment(horizontal="left")

        # Description
        dc = ws.cell(row=row, column=3, value=description)
        dc.font = Font(size=10, name="Calibri")
        dc.fill = fill
        dc.border = _thin()
        dc.alignment = Alignment(horizontal="left")

        # Row count
        if sheet_name in wb.sheetnames:
            row_count = max(0, wb[sheet_name].max_row - 5)
        else:
            row_count = 0
        rc = ws.cell(row=row, column=4, value=row_count if row_count > 0 else "")
        rc.font = Font(size=10, name="Calibri")
        rc.fill = fill
        rc.border = _thin()
        rc.alignment = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 10
    ws.sheet_view.showGridLines = False


def _thin():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def suggested_filename(company: str = "") -> str:
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    slug = company.replace(" ", "_")[:20] if company else "Report"
    return f"FinAnalyzer_{slug}_{ts}.xlsx"
