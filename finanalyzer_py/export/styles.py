from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Palette (matches xlsx-js-style colours in the original app) ──────────────
NAVY    = "1E3A8A"
LBLUE   = "EFF6FF"
DGRAY   = "111827"
TEAL    = "0F766E"
WHITE   = "FFFFFF"
LGRAY   = "F8FAFC"
YELLOW  = "FEF3C7"
BROWN   = "78350F"
BORDER  = "CBD5E1"
RED_BG  = "FEE2E2"
AMBER   = "FEF3C7"
GREEN   = "DCFCE7"
BLUE_BG = "E0E7FF"

# Anomaly / bucket colours
SALES_COL   = "D1FAE5"
PURCH_COL   = "EDE9FE"
EXP_COL     = "FEF3C7"
TDS_COL     = "FECDD3"
GST_COL     = "E0F2FE"
RCM_COL     = "FAE8FF"
BANK_COL    = "E0E7FF"
OTHER_COL   = "F1F5F9"


def _side(style: str = "thin") -> Side:
    return Side(style=style, color=BORDER)


def thin_border() -> Border:
    s = _side("thin")
    return Border(left=s, right=s, top=s, bottom=s)


def medium_border() -> Border:
    s = _side("medium")
    return Border(left=s, right=s, top=s, bottom=s)


def top_medium_border() -> Border:
    return Border(
        left=_side("thin"), right=_side("thin"),
        top=_side("medium"), bottom=_side("thin"),
    )


# ── Fonts ─────────────────────────────────────────────────────────────────────
TITLE_FONT   = Font(bold=True, color=WHITE, size=16, name="Calibri")
META_FONT    = Font(bold=True, color=NAVY, size=10, name="Calibri")
HEADER_FONT  = Font(bold=True, color=WHITE, size=11, name="Calibri")
DATA_FONT    = Font(size=10, name="Calibri")
TOTAL_FONT   = Font(bold=True, color=WHITE, size=10, name="Calibri")
OBS_FONT     = Font(color=BROWN, size=10, name="Calibri")
BOLD_FONT    = Font(bold=True, size=10, name="Calibri")

# ── Fills ─────────────────────────────────────────────────────────────────────
TITLE_FILL   = PatternFill("solid", fgColor=NAVY)
META_FILL    = PatternFill("solid", fgColor=LBLUE)
HEADER_FILL  = PatternFill("solid", fgColor=DGRAY)
ODD_FILL     = PatternFill("solid", fgColor=WHITE)
EVEN_FILL    = PatternFill("solid", fgColor=LGRAY)
TOTAL_FILL   = PatternFill("solid", fgColor=TEAL)
OBS_FILL     = PatternFill("solid", fgColor=YELLOW)
RED_FILL     = PatternFill("solid", fgColor=RED_BG)
GREEN_FILL   = PatternFill("solid", fgColor=GREEN)
BLUE_FILL    = PatternFill("solid", fgColor=BLUE_BG)
AMBER_FILL   = PatternFill("solid", fgColor=AMBER)

def hex_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color.lstrip("#"))

# ── Alignments ────────────────────────────────────────────────────────────────
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Number Formats ────────────────────────────────────────────────────────────
AMOUNT_FMT  = "#,##0.00"
COUNT_FMT   = "#,##0"
PERCENT_FMT = '0.00"%"'
DATE_FMT    = "DD/MM/YYYY"
